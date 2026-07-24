import csv
import json
import logging
import os
import secrets
import sqlite3
import tempfile
import threading
import time
import unicodedata
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from functools import wraps
from io import BytesIO
from xml.sax.saxutils import escape as xml_escape

from dotenv import load_dotenv
from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_wtf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash, generate_password_hash

from werkzeug.exceptions import RequestEntityTooLarge

from db import (
    BASE_DIR,
    DATABASE,
    MIN_PASSWORD_LENGTH,
    BancoInvalido,
    caminho_backup,
    carregar_perfis_mensais,
    clear_setup_token,
    criar_backup,
    ensure_setup_token,
    exportar_snapshot,
    garantir_senha_inicial,
    get_db_connection,
    get_preferencia,
    get_senha_hash,
    importar_banco,
    init_db,
    listar_backups,
    primeiro_mes_perfil,
    restaurar_backup,
    salvar_perfil_mensal,
    set_preferencia,
    set_senha_hash,
    verify_setup_token,
)

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("saude_simples")

app = Flask(__name__)

SECRET_KEY = os.environ.get("SAUDE_SIMPLES_SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError(
        "SAUDE_SIMPLES_SECRET_KEY não definida. Crie um arquivo .env (veja .env.example) "
        "ou gere uma com `python -c \"import secrets; print(secrets.token_hex(32))\"`."
    )
app.secret_key = SECRET_KEY

BOOTSTRAP_PASSWORD_HASH = os.environ.get("SAUDE_SIMPLES_PASSWORD_HASH")

csrf = CSRFProtect(app)

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
# O sistema roda em rede local HTTP por padrão; se colocar TLS na frente,
# ligue SAUDE_SIMPLES_FORCE_HTTPS=true para o cookie só trafegar via HTTPS.
app.config["SESSION_COOKIE_SECURE"] = os.environ.get(
    "SAUDE_SIMPLES_FORCE_HTTPS", ""
).lower() in ("1", "true", "sim", "yes")
# Nenhum formulário legítimo chega perto disso — bloqueia payloads gigantes
# antes de qualquer parsing.
app.config["MAX_CONTENT_LENGTH"] = 1 * 1024 * 1024

# Exceções: importação de banco recebe um arquivo .db real; importação de
# pacientes recebe o CSV do e-SUS (grande em áreas com muitos cadastros).
IMPORT_MAX_BYTES = 200 * 1024 * 1024
CSV_IMPORT_MAX_BYTES = 20 * 1024 * 1024


@app.before_request
def ajustar_limite_de_upload():
    if request.path == "/banco/importar":
        request.max_content_length = IMPORT_MAX_BYTES
    elif request.path == "/pacientes/importar":
        request.max_content_length = CSV_IMPORT_MAX_BYTES


@app.errorhandler(RequestEntityTooLarge)
def arquivo_grande_demais(_error):
    flash("O arquivo enviado excede o tamanho máximo permitido.", "danger")
    if request.path.startswith("/banco"):
        destino = "banco"
    elif request.path.startswith("/pacientes"):
        destino = "importar_pacientes"
    else:
        destino = "index"
    return redirect(url_for(destino))

APP_VERSION = "2.0.0"


def _compute_asset_version():
    """Token de cache-busting dos estáticos: APP_VERSION + mtime mais recente da
    pasta static. Muda automaticamente sempre que qualquer arquivo estático é
    alterado, então não depende de lembrar de subir a versão à mão."""
    latest = 0.0
    static_dir = os.path.join(BASE_DIR, "static")
    for root, _dirs, files in os.walk(static_dir):
        for name in files:
            try:
                latest = max(latest, os.path.getmtime(os.path.join(root, name)))
            except OSError:
                continue
    return f"{APP_VERSION}-{int(latest)}" if latest else APP_VERSION


ASSET_VERSION = _compute_asset_version()


@app.before_request
def prepare_csp_nonce():
    # Um nonce novo por resposta permite scripts/estilos inline autorizados
    # sem liberar execução inline para qualquer conteúdo injetado.
    g.csp_nonce = secrets.token_urlsafe(24)


@app.context_processor
def inject_csp_nonce():
    return {"csp_nonce": g.get("csp_nonce", "")}


# Cache-busting: carimba todo url_for('static', ...) com ?v=ASSET_VERSION.
# Uma versão nova => URLs novas => o navegador baixa os estáticos atuais e
# abandona os da versão anterior. Não exige tocar nos templates.
@app.url_defaults
def _stamp_static_version(endpoint, values):
    if endpoint == "static" and values is not None and "filename" in values:
        values.setdefault("v", ASSET_VERSION)


@app.context_processor
def inject_asset_version():
    return {"asset_version": ASSET_VERSION}


@app.after_request
def set_security_headers(response):
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    if app.config["SESSION_COOKIE_SECURE"]:
        # Servindo sob TLS (run.py liga FORCE_HTTPS): navegadores passam a
        # recusar downgrade para HTTP neste host.
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000")
    nonce = g.get("csp_nonce", "")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        f"script-src 'self' 'nonce-{nonce}'; "
        f"style-src 'self' 'nonce-{nonce}'; "
        # O Alpine (build CSP) alterna visibilidade via style=. A exceção fica
        # restrita a atributos; blocos <style> continuam protegidos pelo nonce.
        "style-src-attr 'unsafe-inline'; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "object-src 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'none'",
    )
    return response

LOGIN_ATTEMPT_LIMIT = 5
LOGIN_ATTEMPT_WINDOW_SECONDS = 60
_login_attempts = {}
# waitress atende com múltiplas threads; o lock evita que duas falhas
# simultâneas do mesmo IP se sobrescrevam e furem o limite.
_login_attempts_lock = threading.Lock()


def registrar_falha_login(client_ip):
    now = time.monotonic()
    with _login_attempts_lock:
        attempts = [
            t for t in _login_attempts.get(client_ip, []) if now - t < LOGIN_ATTEMPT_WINDOW_SECONDS
        ]
        attempts.append(now)
        _login_attempts[client_ip] = attempts


def login_bloqueado(client_ip):
    now = time.monotonic()
    with _login_attempts_lock:
        attempts = [
            t for t in _login_attempts.get(client_ip, []) if now - t < LOGIN_ATTEMPT_WINDOW_SECONDS
        ]
        # Sem tentativas dentro da janela, a chave do IP sai do dict — chamado
        # a cada POST de login/setup/troca de senha, isto mantém a estrutura do
        # tamanho dos IPs ativos, não do total histórico que já bateu à porta.
        if attempts:
            _login_attempts[client_ip] = attempts
        else:
            _login_attempts.pop(client_ip, None)
        return len(attempts) >= LOGIN_ATTEMPT_LIMIT


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("usuario_autenticado"):
            return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
        return view(*args, **kwargs)

    return wrapped


def proxima_url_segura(value):
    if value and value.startswith("/") and not value.startswith("//"):
        return value
    return url_for("index")


def requisicao_e_local():
    # request.remote_addr é o IP de quem conectou direto no socket — não vem de
    # cabeçalho (X-Forwarded-For), então não é falsificável pelo cliente.
    # ATENÇÃO: se este app for colocado detrás de um reverse proxy (nginx, etc.)
    # na mesma máquina, o proxy passa a ser "quem conectou", e todo mundo do lado
    # de fora passaria a aparecer como local. Nesse cenário, configure
    # werkzeug.middleware.proxy_fix.ProxyFix ou desative esta rota.
    return request.remote_addr in ("127.0.0.1", "::1")

TIPOS_IMOVEL_OPCOES = [
    {"codigo": "domicilio", "label": "Domicílio"},
    {"codigo": "apartamento", "label": "Apartamento"},
    {"codigo": "loja", "label": "Loja"},
    {"codigo": "lava_jato", "label": "Lava Jato"},
    {"codigo": "terreno_baldio", "label": "Terreno Baldio"},
    {"codigo": "escola", "label": "Escola"},
    {"codigo": "igreja", "label": "Igreja"},
    {"codigo": "pizzaria", "label": "Pizzaria"},
    {"codigo": "hamburgueria", "label": "Hamburgueria"},
    {"codigo": "em_construcao", "label": "Em construção"},
]
TIPOS_IMOVEL_POR_CODIGO = {opcao["codigo"]: opcao["label"] for opcao in TIPOS_IMOVEL_OPCOES}

# Tipos onde mora gente: sem morador cadastrado, contam como "vazia" no painel.
TIPOS_IMOVEL_RESIDENCIAIS = {"domicilio", "apartamento"}


def normalizar_tipo_imovel(value):
    value = str(value or "").strip()
    return value if value in TIPOS_IMOVEL_POR_CODIGO else "domicilio"


def tipo_imovel_label(value):
    return TIPOS_IMOVEL_POR_CODIGO.get(str(value or "").strip(), "Domicílio")


def tipo_imovel_de(casa):
    """Lê o tipo de uma linha de casa tolerando bancos antigos importados
    (coluna pode não existir até o init_db rodar) e valores nulos."""
    try:
        return normalizar_tipo_imovel(casa["tipo_imovel"])
    except (IndexError, KeyError):
        return "domicilio"


CONDICOES_SAUDE_OPCOES = [
    {"codigo": "gestante", "label": "Está gestante"},
    {"codigo": "abaixo_peso", "label": "Abaixo do peso"},
    {"codigo": "peso_adequado", "label": "Peso adequado"},
    {"codigo": "acima_peso", "label": "Acima do peso"},
    {"codigo": "fumante", "label": "Está fumante"},
    {"codigo": "alcool", "label": "Faz uso de álcool"},
    {"codigo": "outras_drogas", "label": "Faz uso de outras drogas"},
    {"codigo": "hipertensao", "label": "Tem hipertensão arterial"},
    {"codigo": "diabetes", "label": "Tem diabetes"},
    {"codigo": "avc_derrame", "label": "Teve AVC/derrame"},
    {"codigo": "infarto", "label": "Teve infarto"},
    {"codigo": "doenca_cardiaca", "label": "Tem doença cardíaca/do coração"},
    {"codigo": "problemas_rins", "label": "Tem ou teve problemas nos rins"},
    {"codigo": "doenca_respiratoria", "label": "Tem doença respiratória/no pulmão"},
    {"codigo": "asma", "label": "Asma"},
    {"codigo": "dpoc_enfisema", "label": "DPOC/Enfisema"},
    {"codigo": "hanseniase", "label": "Hanseníase"},
    {"codigo": "tuberculose", "label": "Tuberculose"},
    {"codigo": "cancer", "label": "Tem ou teve câncer"},
    {"codigo": "internacao_12_meses", "label": "Teve internação nos últimos 12 meses"},
    {"codigo": "saude_mental", "label": "Diagnóstico de problema de saúde mental"},
    {"codigo": "acamado", "label": "Está acamado"},
    {"codigo": "domiciliado", "label": "Está domiciliado"},
    {"codigo": "deficiencia", "label": "Pessoa com deficiência"},
    {"codigo": "plantas_medicinais", "label": "Usa plantas medicinais"},
    {"codigo": "praticas_integrativas", "label": "Usa práticas integrativas e complementares"},
    {"codigo": "bolsa_familia", "label": "Recebe Bolsa Família"},
    {"codigo": "outras_condicoes", "label": "Outras condições de saúde"},
]
CONDICOES_SAUDE = [opcao["label"] for opcao in CONDICOES_SAUDE_OPCOES]

# Situação cadastral do paciente. Quem se muda ou falece não é excluído: o
# cadastro é preservado com o motivo da saída, fora das contagens do território.
STATUS_PACIENTE_OPCOES = [
    {"codigo": "ativo", "label": "Ativo"},
    {"codigo": "mudou_se", "label": "Mudou-se"},
    {"codigo": "fora_de_area", "label": "Fora de Área"},
    {"codigo": "obito", "label": "Óbito"},
]
STATUS_PACIENTE_POR_CODIGO = {opcao["codigo"]: opcao["label"] for opcao in STATUS_PACIENTE_OPCOES}

# Nos JOINs/WHEREs: COALESCE cobre bancos importados onde a coluna possa
# existir sem NOT NULL — NULL conta como ativo, igual ao legado.
SQL_PACIENTE_ATIVO = "COALESCE(pacientes.status, 'ativo') = 'ativo'"


def normalizar_status_paciente(value):
    value = str(value or "").strip()
    return value if value in STATUS_PACIENTE_POR_CODIGO else "ativo"


def status_paciente_label(value):
    return STATUS_PACIENTE_POR_CODIGO.get(str(value or "").strip(), "Ativo")


def status_paciente_de(paciente):
    """Lê o status de uma linha de paciente tolerando bancos antigos importados
    (coluna pode não existir até o init_db rodar) e valores nulos."""
    try:
        return normalizar_status_paciente(paciente["status"])
    except (IndexError, KeyError):
        return "ativo"


def get_next_house_number(quadra_id=None):
    conn = get_db_connection()
    if quadra_id:
        row = conn.execute(
            "SELECT COALESCE(MAX(numero_casa), 0) + 1 AS proximo FROM casas WHERE quadra_id = ?",
            (quadra_id,),
        ).fetchone()
    else:
        row = conn.execute("SELECT COALESCE(MAX(numero_casa), 0) + 1 AS proximo FROM casas").fetchone()
    conn.close()
    return row["proximo"]


def get_next_block_number():
    conn = get_db_connection()
    row = conn.execute("SELECT COALESCE(MAX(numero_quadra), 0) + 1 AS proximo FROM quadras").fetchone()
    conn.close()
    return row["proximo"]


def get_quadras():
    conn = get_db_connection()
    quadras = conn.execute("SELECT * FROM quadras ORDER BY numero_quadra, id").fetchall()
    conn.close()
    return quadras


def parse_optional_int(value):
    value = str(value or "").strip()
    return int(value) if value else None


def parse_positive_int(value, field_name, required=True):
    value = str(value or "").strip()
    if not value:
        if required:
            return None, f"Informe {field_name}."
        return None, None

    try:
        number = int(value)
    except ValueError:
        return None, f"{field_name.capitalize()} deve ser um número válido."

    if number < 1:
        return None, f"{field_name.capitalize()} deve ser maior que zero."
    return number, None


def quadra_exists(quadra_id):
    if quadra_id is None:
        return True

    conn = get_db_connection()
    exists = conn.execute("SELECT 1 FROM quadras WHERE id = ?", (quadra_id,)).fetchone() is not None
    conn.close()
    return exists


def pdf_text(value, style):
    return Paragraph(xml_escape(str(value or "")), style)


def pdf_count(value, style):
    return Paragraph(xml_escape(str(value if value not in (None, "") else 0)), style)


def pdf_image(path, max_width, max_height):
    image_source = path
    try:
        with PILImage.open(path) as source:
            source.thumbnail((1400, 700))
            optimized = BytesIO()
            source.convert("RGB").save(optimized, format="JPEG", quality=82, optimize=True)
            optimized.seek(0)
            image_source = optimized
    except (OSError, ValueError) as exc:
        logger.warning("Falha ao otimizar imagem do relatório %s: %s", path, exc)
        image_source = path

    image = Image(image_source)
    ratio = min(max_width / image.imageWidth, max_height / image.imageHeight)
    image.drawWidth = image.imageWidth * ratio
    image.drawHeight = image.imageHeight * ratio
    image.hAlign = "CENTER"
    return image


def formatar_cpf_ou_cns(value):
    digits = "".join(char for char in str(value or "") if char.isdigit())[:15]

    if len(digits) <= 11:
        if len(digits) <= 3:
            return digits
        if len(digits) <= 6:
            return f"{digits[:3]}.{digits[3:]}"
        if len(digits) <= 9:
            return f"{digits[:3]}.{digits[3:6]}.{digits[6:]}"
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"

    parts = [digits[:3], digits[3:7], digits[7:11], digits[11:15]]
    return " ".join(part for part in parts if part)


def formatar_data_br(value):
    if not value:
        return ""

    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return value


def formatar_telefone(value):
    digits = "".join(char for char in str(value or "") if char.isdigit())[:11]

    if len(digits) <= 2:
        return digits
    if len(digits) <= 6:
        return f"({digits[:2]}) {digits[2:]}"
    if len(digits) <= 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"


def whatsapp_link(value):
    digits = "".join(char for char in str(value or "") if char.isdigit())
    if not digits:
        return ""
    if len(digits) in (10, 11):
        digits = f"55{digits}"
    return f"https://wa.me/{digits}"


def normalizar_condicoes(values):
    selected = []
    for value in values:
        value = value.strip()
        codigo = condicao_codigo(value)
        if codigo and codigo not in selected:
            selected.append(codigo)
    return "\n".join(selected)


def listar_condicoes(value):
    return [condicao_label(item) for item in str(value or "").splitlines() if item]


def listar_condicao_codigos(value):
    return [condicao_codigo(item) for item in str(value or "").splitlines() if condicao_codigo(item)]


def paciente_tem_condicoes(paciente, condicoes):
    if not condicoes:
        return True
    return bool(set(listar_condicao_codigos(paciente["condicoes_saude"])) & set(condicoes))


def texto_normalizado(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in value if not unicodedata.combining(char)).lower()


CONDICOES_POR_CODIGO = {opcao["codigo"]: opcao["label"] for opcao in CONDICOES_SAUDE_OPCOES}
CONDICOES_CODIGO_POR_LABEL = {
    texto_normalizado(opcao["label"]): opcao["codigo"] for opcao in CONDICOES_SAUDE_OPCOES
}


def condicao_codigo(value):
    value = str(value or "").strip()
    if value in CONDICOES_POR_CODIGO:
        return value
    return CONDICOES_CODIGO_POR_LABEL.get(texto_normalizado(value), "")


def condicao_label(value):
    value = str(value or "").strip()
    if value in CONDICOES_POR_CODIGO:
        return CONDICOES_POR_CODIGO[value]
    codigo = CONDICOES_CODIGO_POR_LABEL.get(texto_normalizado(value))
    return CONDICOES_POR_CODIGO.get(codigo, value)


def apenas_digitos(value):
    return "".join(char for char in str(value or "") if char.isdigit())


def paciente_corresponde_busca(paciente, busca):
    busca_normalizada = texto_normalizado(busca).strip()
    busca_digitos = apenas_digitos(busca)
    nome_normalizado = texto_normalizado(paciente["nome"])
    documento_digitos = apenas_digitos(paciente["cpf"])

    if busca_digitos and busca_digitos in documento_digitos:
        return True

    if not busca_normalizada:
        return False

    if busca_normalizada in nome_normalizado:
        return True

    termos_busca = busca_normalizada.split()
    termos_nome = nome_normalizado.split()

    if termos_busca and all(
        any(
            termo in termo_nome or SequenceMatcher(None, termo, termo_nome).ratio() >= 0.62
            for termo_nome in termos_nome
        )
        for termo in termos_busca
    ):
        return True

    return SequenceMatcher(None, busca_normalizada, nome_normalizado).ratio() >= 0.58


def calcular_idade(data_nascimento):
    if not data_nascimento:
        return None

    try:
        nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d").date()
    except ValueError:
        return None

    hoje = datetime.now().date()
    idade = hoje.year - nascimento.year
    if (hoje.month, hoje.day) < (nascimento.month, nascimento.day):
        idade -= 1
    return idade


# ---------------------------------------------------------------------------
# Grupos demográficos — fonte única de verdade
#
# Os mesmos predicados alimentam os cards do painel, as stats do PDF e os
# recortes de exportação. Mudou a regra (ex.: idade de corte), muda em um
# lugar só.
# ---------------------------------------------------------------------------
def paciente_e_crianca(paciente):
    # `idade or 999` seria armadilha: 0 (menos de 1 ano) é falsy — recém-
    # -nascido tem que contar como criança.
    idade = calcular_idade(paciente["data_nascimento"])
    return idade is not None and idade < 12


def paciente_e_crianca_0_2(paciente):
    idade = calcular_idade(paciente["data_nascimento"])
    return idade is not None and idade <= 2


def paciente_e_idoso(paciente):
    return (calcular_idade(paciente["data_nascimento"]) or 0) >= 60


def paciente_e_gestante(paciente):
    codigos = listar_condicao_codigos(paciente["condicoes_saude"])
    if "gestante" in codigos:
        return True
    # Bancos antigos guardavam o texto livre da condição.
    return any(
        "gestante" in texto_normalizado(condicao)
        for condicao in listar_condicoes(paciente["condicoes_saude"])
    )


def paciente_e_homem(paciente):
    return str(paciente["sexo"] or "").lower() == "masculino"


def paciente_e_mulher(paciente):
    return str(paciente["sexo"] or "").lower() == "feminino"


# "Pacientes (todos)" do painel é o estado sem grupo marcado — não é um grupo.
GRUPOS_DEMOGRAFICOS_OPCOES = [
    {"codigo": "criancas_0_2", "label": "Crianças 0 a 2 anos"},
    {"codigo": "criancas", "label": "Crianças (até 11 anos)"},
    {"codigo": "idosos", "label": "Idosos 60+"},
    {"codigo": "gestantes", "label": "Gestantes"},
    {"codigo": "homens", "label": "Homens"},
    {"codigo": "mulheres", "label": "Mulheres"},
]
GRUPOS_POR_CODIGO = {opcao["codigo"]: opcao["label"] for opcao in GRUPOS_DEMOGRAFICOS_OPCOES}
GRUPOS_PREDICADOS = {
    "criancas_0_2": paciente_e_crianca_0_2,
    "criancas": paciente_e_crianca,
    "idosos": paciente_e_idoso,
    "gestantes": paciente_e_gestante,
    "homens": paciente_e_homem,
    "mulheres": paciente_e_mulher,
}


def ler_codigos_grupos(values):
    """Whitelist dos grupos vindos da query string, preservando a ordem."""
    codigos = []
    for value in values:
        value = str(value or "").strip()
        if value in GRUPOS_PREDICADOS and value not in codigos:
            codigos.append(value)
    return codigos


def paciente_em_grupos(paciente, grupos):
    """União: sem grupo marcado todo mundo passa; com grupos, basta pertencer
    a um deles."""
    if not grupos:
        return True
    return any(GRUPOS_PREDICADOS[grupo](paciente) for grupo in grupos)


def rotulo_grupos(grupos):
    labels = [GRUPOS_POR_CODIGO[grupo] for grupo in grupos]
    if len(labels) <= 1:
        return "".join(labels)
    return ", ".join(labels[:-1]) + " e " + labels[-1]


def build_dashboard_stats(casas, pacientes):
    total_casas = len(casas)
    # "Vazia" é um achado de campo apenas para imóvel residencial — loja,
    # escola ou terreno baldio sem morador é o esperado, não um alerta.
    casas_vazias = sum(
        1
        for casa in casas
        if casa["total_pacientes"] == 0 and tipo_imovel_de(casa) in TIPOS_IMOVEL_RESIDENCIAIS
    )
    total_pacientes = len(pacientes)
    # Mesmos predicados dos recortes de exportação — uma única régua.
    criancas = sum(1 for paciente in pacientes if paciente_e_crianca(paciente))
    idosos = sum(1 for paciente in pacientes if paciente_e_idoso(paciente))
    gestantes = sum(1 for paciente in pacientes if paciente_e_gestante(paciente))
    homens = sum(1 for paciente in pacientes if paciente_e_homem(paciente))
    mulheres = sum(1 for paciente in pacientes if paciente_e_mulher(paciente))
    comorbidades = {condicao: 0 for condicao in CONDICOES_SAUDE}

    for paciente in pacientes:
        for codigo in listar_condicao_codigos(paciente["condicoes_saude"]):
            condicao = condicao_label(codigo)
            comorbidades[condicao] = comorbidades.get(condicao, 0) + 1

    return {
        "total_casas": total_casas,
        "casas_vazias": casas_vazias,
        "total_pacientes": total_pacientes,
        "criancas": criancas,
        "idosos": idosos,
        "gestantes": gestantes,
        "homens": homens,
        "mulheres": mulheres,
        "comorbidades": sorted(comorbidades.items(), key=lambda item: item[0]),
    }


def ler_codigos_condicoes(values):
    codigos = []
    for condicao in values:
        codigo = condicao_codigo(condicao)
        if codigo and codigo not in codigos:
            codigos.append(codigo)
    return codigos


def carregar_dados_relatorio(condicoes_selecionadas=None, filtro_sem_selecao=False, grupos_selecionados=None):
    condicoes_selecionadas = condicoes_selecionadas or []
    grupos_selecionados = grupos_selecionados or []
    conn = get_db_connection()
    casas = conn.execute(
        f"""
        SELECT casas.*, quadras.numero_quadra, COUNT(pacientes.id) AS total_pacientes
        FROM casas
        LEFT JOIN quadras ON quadras.id = casas.quadra_id
        LEFT JOIN pacientes ON pacientes.casa_id = casas.id AND {SQL_PACIENTE_ATIVO}
        GROUP BY casas.id
        ORDER BY quadras.numero_quadra IS NULL, quadras.numero_quadra, casas.numero_casa, casas.id
        """
    ).fetchall()
    pacientes = conn.execute(
        f"SELECT * FROM pacientes WHERE {SQL_PACIENTE_ATIVO} ORDER BY nome"
    ).fetchall()
    conn.close()

    if condicoes_selecionadas or grupos_selecionados:
        # Grupos entre si: UNIÃO. Grupos × comorbidades: INTERSEÇÃO.
        pacientes = [
            paciente
            for paciente in pacientes
            if paciente_tem_condicoes(paciente, condicoes_selecionadas)
            and paciente_em_grupos(paciente, grupos_selecionados)
        ]
        total_por_casa = {}
        for paciente in pacientes:
            total_por_casa[paciente["casa_id"]] = total_por_casa.get(paciente["casa_id"], 0) + 1

        # Só casas com pacientes no recorte entram no relatório.
        casas = [
            {**dict(casa), "total_pacientes": total_por_casa[casa["id"]]}
            for casa in casas
            if casa["id"] in total_por_casa
        ]
    elif filtro_sem_selecao:
        casas = []
        pacientes = []

    return casas, pacientes


# ---------------------------------------------------------------------------
# Perfil epidemiológico mensal
#
# O retrato do mês corrente é recalculado do banco e gravado (upsert) no
# primeiro acesso autenticado de cada dia e a cada exportação do perfil.
# Quando o mês vira, o último retrato gravado congela como histórico — a
# série mensal se constrói sozinha, sem rotina manual de fechamento. Mês em
# que o sistema não foi usado fica sem retrato e aparece como "-----".
# ---------------------------------------------------------------------------
MESES_ABREVIADOS = ("JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ")

# Linhas do relatório: (rótulo, chave no retrato, "fm" separa F/M, "span" ocupa
# as duas colunas do mês. Ordem e rótulos seguem o modelo oficial do perfil.
PERFIL_LINHAS = [
    ("Pessoas cadastradas", "pessoas", "fm"),
    ("Hipertensos", "hipertensos", "fm"),
    ("Diabéticos", "diabeticos", "fm"),
    ("Idosos", "idosos", "fm"),
    ("Gestantes", "gestantes", "span"),
    ("Crianças 0 a 2 anos", "criancas_0_2", "fm"),
    ("Acamados", "acamados", "fm"),
    ("Domiciliados", "domiciliados", "fm"),
    ("Obesos", "obesos", "fm"),
    ("Saúde Mental", "saude_mental", "fm"),
    ("Deficientes", "deficientes", "fm"),
    ("Tabagistas", "tabagistas", "fm"),
    ("Bolsa Família", "bolsa_familia", "fm"),
]


def calcular_perfil_epidemiologico():
    """Retrato atual dos indicadores do território, separados por sexo."""
    casas, pacientes = carregar_dados_relatorio()

    def por_sexo(predicado):
        return {
            "f": sum(1 for p in pacientes if paciente_e_mulher(p) and predicado(p)),
            "m": sum(1 for p in pacientes if paciente_e_homem(p) and predicado(p)),
        }

    def por_condicao(codigo):
        return por_sexo(lambda p: codigo in listar_condicao_codigos(p["condicoes_saude"]))

    return {
        # Família = casa com pelo menos um morador ativo (casa vazia não conta).
        "familias": sum(1 for casa in casas if casa["total_pacientes"] > 0),
        "pessoas": por_sexo(lambda p: True),
        "hipertensos": por_condicao("hipertensao"),
        "diabeticos": por_condicao("diabetes"),
        "idosos": por_sexo(paciente_e_idoso),
        "gestantes": sum(1 for p in pacientes if paciente_e_gestante(p)),
        "criancas_0_2": por_sexo(paciente_e_crianca_0_2),
        "acamados": por_condicao("acamado"),
        "domiciliados": por_condicao("domiciliado"),
        "obesos": por_condicao("acima_peso"),
        "saude_mental": por_condicao("saude_mental"),
        "deficientes": por_condicao("deficiencia"),
        "tabagistas": por_condicao("fumante"),
        "bolsa_familia": por_condicao("bolsa_familia"),
    }


def registrar_perfil_mensal():
    """Grava (upsert) o retrato do mês corrente. Retorna o retrato."""
    perfil = calcular_perfil_epidemiologico()
    ano_mes = datetime.now().strftime("%Y-%m")
    salvar_perfil_mensal(ano_mes, json.dumps(perfil, ensure_ascii=False))
    return perfil


# Dia (AAAA-MM-DD) em que o retrato já foi registrado neste processo — o hook
# abaixo roda no máximo uma vez por dia. Corrida entre threads é inofensiva:
# o pior caso é um upsert repetido com o mesmo conteúdo.
_perfil_registrado_dia = None


@app.before_request
def registrar_perfil_do_dia():
    global _perfil_registrado_dia
    hoje = datetime.now().strftime("%Y-%m-%d")
    if _perfil_registrado_dia == hoje:
        return
    if request.endpoint == "static" or not session.get("usuario_autenticado"):
        return
    _perfil_registrado_dia = hoje
    try:
        registrar_perfil_mensal()
    except sqlite3.Error:
        # Nunca derruba a navegação por causa do retrato — tenta de novo amanhã.
        logger.exception("Falha ao registrar o perfil mensal automático")
        _perfil_registrado_dia = None


# Teto de sanidade da série (10 anos). Não é janela: nenhum mês registrado
# dentro dele é descartado — só protege o documento de um banco com ano_mes
# absurdo (ex.: editado à mão) virar um PDF de milhares de colunas.
PERFIL_MAXIMO_MESES = 120

# O PDF é o entregável mensal: mostra os últimos 12 meses, sempre com um ano
# de contexto. O mês que sai do PDF permanece no banco e sai integral no XLSX
# — a janela é de exibição, nunca de retenção.
PERFIL_MESES_NO_PDF = 12


def meses_da_serie_perfil(referencia=None):
    """Do primeiro mês com retrato registrado até o mês corrente, contínuo.

    A série nasce no mês em que o sistema começou a ser alimentado — não há
    janela fixa nem data cravada em código: o histórico só cresce, mês a mês.
    Sem nenhum retrato ainda, a série é só o mês corrente. Buracos no meio
    (mês sem uso) permanecem na série e saem como "-----" no documento.
    """
    referencia = referencia or datetime.now()
    atual = f"{referencia.year:04d}-{referencia.month:02d}"
    inicio = primeiro_mes_perfil() or atual
    if inicio > atual:
        inicio = atual  # relógio ajustado para trás não pode quebrar a série

    ano, mes = int(inicio[:4]), int(inicio[5:7])
    total = (referencia.year - ano) * 12 + (referencia.month - mes) + 1
    if total > PERFIL_MAXIMO_MESES:
        # Acima do teto, ficam os meses mais RECENTES — o corte é no passado.
        avanco = total - PERFIL_MAXIMO_MESES
        mes += avanco
        ano += (mes - 1) // 12
        mes = (mes - 1) % 12 + 1
        total = PERFIL_MAXIMO_MESES

    meses = []
    for _ in range(total):
        meses.append(f"{ano:04d}-{mes:02d}")
        mes += 1
        if mes == 13:
            ano, mes = ano + 1, 1
    return meses


def rotulo_mes(ano_mes):
    ano, mes = ano_mes.split("-")
    return f"{MESES_ABREVIADOS[int(mes) - 1]}/{ano}"


def montar_serie_perfis(meses, perfil_atual):
    """Retratos mês a mês, com carência herdada: mês sem registro recebe o
    último retrato anterior — sistema parado significa cadastro inalterado,
    não dado perdido. O relatório sai entregável em qualquer mês.

    O mês corrente é sempre o `perfil_atual` (calculado ao vivo). "-----" só
    resta quando não há retrato nenhum antes do mês (série corrompida) ou
    quando um indicador não existia no retrato da época.
    """
    salvos = carregar_perfis_mensais(meses[:-1])
    perfis = {}
    ultimo = None
    for ano_mes in meses[:-1]:
        dados = None
        if ano_mes in salvos:
            try:
                dados = json.loads(salvos[ano_mes])
            except ValueError:
                dados = None  # registro ilegível cai na herança, não quebra
        perfis[ano_mes] = dados if dados is not None else ultimo
        if dados is not None:
            ultimo = dados
    perfis[meses[-1]] = perfil_atual
    return perfis


def formatar_tamanho(num_bytes):
    num_bytes = int(num_bytes or 0)
    for unidade in ("B", "KB", "MB", "GB"):
        if num_bytes < 1024 or unidade == "GB":
            if unidade == "B":
                return f"{num_bytes} {unidade}"
            return f"{num_bytes:.1f} {unidade}"
        num_bytes /= 1024
    return f"{num_bytes:.1f} GB"


app.add_template_filter(formatar_tamanho, "tamanho")
app.add_template_filter(tipo_imovel_label, "tipo_imovel")
app.add_template_filter(formatar_cpf_ou_cns, "cpf_cns")
app.add_template_filter(formatar_data_br, "data_br")
app.add_template_filter(formatar_telefone, "telefone")
app.add_template_filter(whatsapp_link, "whatsapp")
app.add_template_filter(listar_condicoes, "condicoes_lista")
app.add_template_filter(listar_condicao_codigos, "condicoes_codigos")
app.add_template_filter(status_paciente_label, "status_paciente")
app.add_template_filter(status_paciente_de, "status_codigo")


@app.context_processor
def inject_options():
    return {
        "opcoes_condicoes_saude": CONDICOES_SAUDE_OPCOES,
        "opcoes_tipos_imovel": TIPOS_IMOVEL_OPCOES,
        "tipos_imovel_residenciais": TIPOS_IMOVEL_RESIDENCIAIS,
        "opcoes_status_paciente": STATUS_PACIENTE_OPCOES,
    }


@app.errorhandler(404)
def pagina_nao_encontrada(_error):
    return render_template("erro_404.html"), 404


def paciente_com_documento(documento_digitos, excluir_id=None):
    """Paciente existente com o mesmo CPF/CNS (comparação por dígitos, já que
    o campo é armazenado formatado). Garante que um documento identifique um
    único paciente — mesma regra da importação do e-SUS."""
    if not documento_digitos:
        return None

    conn = get_db_connection()
    rows = conn.execute("SELECT id, nome, cpf FROM pacientes").fetchall()
    conn.close()
    for row in rows:
        if excluir_id is not None and row["id"] == excluir_id:
            continue
        if apenas_digitos(row["cpf"]) == documento_digitos:
            return row
    return None


def get_casas_para_transferencia(excluir_casa_id=None):
    """Casas candidatas a destino de transferência, na mesma ordem do painel."""
    conn = get_db_connection()
    casas = conn.execute(
        """
        SELECT casas.id, casas.numero_casa, casas.endereco, quadras.numero_quadra
        FROM casas
        LEFT JOIN quadras ON quadras.id = casas.quadra_id
        ORDER BY quadras.numero_quadra IS NULL, quadras.numero_quadra, casas.numero_casa, casas.id
        """
    ).fetchall()
    conn.close()
    if excluir_casa_id is None:
        return casas
    return [casa for casa in casas if casa["id"] != excluir_casa_id]


def inserir_paciente_do_form(casa_id, nome):
    """INSERT único dos dois fluxos de cadastro (pela casa e pela lista):
    os campos do form são normalizados num lugar só."""
    conn = get_db_connection()
    conn.execute(
        """
        INSERT INTO pacientes (
            casa_id, nome, cpf, telefone, data_nascimento, sexo, nome_pai, nome_mae,
            condicoes_saude, observacao
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            casa_id,
            nome,
            formatar_cpf_ou_cns(request.form.get("cpf", "")),
            formatar_telefone(request.form.get("telefone", "")),
            request.form.get("data_nascimento", "").strip(),
            request.form.get("sexo", "").strip(),
            request.form.get("nome_pai", "").strip(),
            request.form.get("nome_mae", "").strip(),
            normalizar_condicoes(request.form.getlist("condicoes_saude")),
            request.form.get("observacao", "").strip(),
        ),
    )
    conn.commit()
    conn.close()


def redirect_apos_acao_no_paciente(casa_id):
    """Volta para a casa do paciente — ou, sem casa vinculada (ex.: importado
    do e-SUS ainda não alocado), para a lista de pacientes."""
    if casa_id:
        return redirect(url_for("detalhes_casa", casa_id=casa_id))
    return redirect(url_for("listar_pacientes"))


def get_house_or_404(casa_id):
    conn = get_db_connection()
    casa = conn.execute(
        """
        SELECT casas.*, quadras.numero_quadra
        FROM casas
        LEFT JOIN quadras ON quadras.id = casas.quadra_id
        WHERE casas.id = ?
        """,
        (casa_id,),
    ).fetchone()
    conn.close()
    if casa is None:
        flash("Casa não encontrada.", "warning")
        return None
    return casa


@app.route("/setup", methods=["GET", "POST"])
def setup():
    # Uso único e irreversível: com senha configurada, /setup redireciona
    # para o login para sempre.
    if get_senha_hash():
        return redirect(url_for("login"))

    if request.method == "POST":
        client_ip = request.remote_addr or "desconhecido"
        if login_bloqueado(client_ip):
            flash("Muitas tentativas. Aguarde um minuto antes de tentar novamente.", "danger")
            return render_template("setup.html"), 429

        codigo = request.form.get("codigo", "")
        nova_senha = request.form.get("nova_senha", "")
        confirmar_senha = request.form.get("confirmar_senha", "")

        if len(nova_senha) < MIN_PASSWORD_LENGTH:
            flash(f"A senha deve ter pelo menos {MIN_PASSWORD_LENGTH} caracteres.", "danger")
            return render_template("setup.html")

        if nova_senha != confirmar_senha:
            flash("A confirmação não corresponde à senha.", "danger")
            return render_template("setup.html")

        if not verify_setup_token(codigo):
            registrar_falha_login(client_ip)
            logger.warning("Código de configuração inicial inválido a partir de %s", client_ip)
            flash("Código de segurança inválido. Confira o valor exibido no terminal do servidor.", "danger")
            return render_template("setup.html")

        set_senha_hash(generate_password_hash(nova_senha))
        clear_setup_token()
        logger.info("Senha inicial definida via /setup a partir de %s", client_ip)
        flash("Senha definida com sucesso. Entre no sistema.", "success")
        return redirect(url_for("login"))

    # Garante que o código exista (e esteja no terminal/instance) enquanto a
    # tela estiver aberta — mesmo que o servidor tenha reiniciado no meio.
    ensure_setup_token()
    return render_template("setup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    # Instalação recém-criada: não existe senha para conferir — o assistente
    # de primeiro acesso assume.
    if get_senha_hash() is None:
        return redirect(url_for("setup"))

    if session.get("usuario_autenticado"):
        return redirect(url_for("index"))

    if request.method == "POST":
        client_ip = request.remote_addr or "desconhecido"
        if login_bloqueado(client_ip):
            flash("Muitas tentativas. Aguarde um minuto antes de tentar novamente.", "danger")
            return render_template("login.html", recuperacao_local=requisicao_e_local()), 429

        senha = request.form.get("senha", "")

        if check_password_hash(get_senha_hash(), senha):
            session.clear()
            session.permanent = True
            session["usuario_autenticado"] = True
            logger.info("Login bem-sucedido a partir de %s", client_ip)
            return redirect(proxima_url_segura(request.args.get("next")))

        registrar_falha_login(client_ip)
        logger.warning("Tentativa de login falhou a partir de %s", client_ip)
        flash("Senha inválida.", "danger")

    return render_template("login.html", recuperacao_local=requisicao_e_local())


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/recuperar-senha", methods=["GET", "POST"])
def recuperar_senha():
    if not requisicao_e_local():
        flash("A recuperação de senha só está disponível acessando o sistema na própria máquina onde ele roda.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        nova_senha = request.form.get("nova_senha", "")
        confirmar_senha = request.form.get("confirmar_senha", "")

        if len(nova_senha) < MIN_PASSWORD_LENGTH:
            flash(f"A nova senha deve ter pelo menos {MIN_PASSWORD_LENGTH} caracteres.", "danger")
            return render_template("recuperar_senha.html")

        if nova_senha != confirmar_senha:
            flash("A confirmação não corresponde à nova senha.", "danger")
            return render_template("recuperar_senha.html")

        try:
            criar_backup("antes_recuperar_senha")
        except (sqlite3.Error, OSError) as exc:
            logger.error("Falha ao criar backup antes de recuperar senha: %s", exc)
            flash("Não foi possível criar backup de segurança. Tente novamente.", "danger")
            return render_template("recuperar_senha.html")

        set_senha_hash(generate_password_hash(nova_senha))
        session.clear()
        logger.warning("Senha redefinida via recuperação local (acesso 127.0.0.1).")
        flash("Senha redefinida com sucesso. Faça login com a nova senha.", "success")
        return redirect(url_for("login"))

    return render_template("recuperar_senha.html")


@app.route("/conta/senha", methods=["GET", "POST"])
@login_required
def alterar_senha():
    if request.method == "POST":
        client_ip = request.remote_addr or "desconhecido"
        if login_bloqueado(client_ip):
            flash("Muitas tentativas. Aguarde um minuto antes de tentar novamente.", "danger")
            return render_template("alterar_senha.html"), 429

        senha_atual = request.form.get("senha_atual", "")
        nova_senha = request.form.get("nova_senha", "")
        confirmar_senha = request.form.get("confirmar_senha", "")

        if not check_password_hash(get_senha_hash(), senha_atual):
            registrar_falha_login(client_ip)
            logger.warning("Tentativa de troca de senha com senha atual incorreta a partir de %s", client_ip)
            flash("Senha atual incorreta.", "danger")
            return render_template("alterar_senha.html")

        if len(nova_senha) < MIN_PASSWORD_LENGTH:
            flash(f"A nova senha deve ter pelo menos {MIN_PASSWORD_LENGTH} caracteres.", "danger")
            return render_template("alterar_senha.html")

        if nova_senha != confirmar_senha:
            flash("A confirmação não corresponde à nova senha.", "danger")
            return render_template("alterar_senha.html")

        if check_password_hash(get_senha_hash(), nova_senha):
            flash("A nova senha deve ser diferente da senha atual.", "danger")
            return render_template("alterar_senha.html")

        set_senha_hash(generate_password_hash(nova_senha))
        logger.info("Senha alterada com sucesso a partir de %s", client_ip)
        flash("Senha alterada com sucesso.", "success")
        return redirect(url_for("index"))

    return render_template("alterar_senha.html")


@app.route("/")
@login_required
def index():
    busca = request.args.get("busca", "").strip()
    # Ordenação da lista de casas pela numeração: crescente (padrão) ou decrescente.
    ordem = request.args.get("ordem", "asc").strip().lower()
    if ordem not in ("asc", "desc"):
        ordem = "asc"
    direcao_numero = "DESC" if ordem == "desc" else "ASC"
    conn = get_db_connection()
    quadras = conn.execute(
        """
        SELECT quadras.*, COUNT(casas.id) AS total_casas
        FROM quadras
        LEFT JOIN casas ON casas.quadra_id = quadras.id
        GROUP BY quadras.id
        ORDER BY quadras.numero_quadra, quadras.id
        """
    ).fetchall()
    casas = conn.execute(
        f"""
        SELECT casas.*, quadras.numero_quadra, COUNT(pacientes.id) AS total_pacientes
        FROM casas
        LEFT JOIN quadras ON quadras.id = casas.quadra_id
        LEFT JOIN pacientes ON pacientes.casa_id = casas.id AND {SQL_PACIENTE_ATIVO}
        GROUP BY casas.id
        ORDER BY casas.numero_casa IS NULL, casas.numero_casa {direcao_numero},
                 quadras.numero_quadra IS NULL, quadras.numero_quadra, casas.id
        """
    ).fetchall()
    pacientes = conn.execute(
        f"SELECT sexo, data_nascimento, condicoes_saude FROM pacientes WHERE {SQL_PACIENTE_ATIVO}"
    ).fetchall()
    pacientes_busca = []
    if busca:
        todos_pacientes = conn.execute(
            """
            SELECT pacientes.*, casas.numero_casa, casas.endereco, quadras.numero_quadra
            FROM pacientes
            LEFT JOIN casas ON casas.id = pacientes.casa_id
            LEFT JOIN quadras ON quadras.id = casas.quadra_id
            ORDER BY pacientes.nome
            """
        ).fetchall()
        pacientes_busca = [
            paciente for paciente in todos_pacientes if paciente_corresponde_busca(paciente, busca)
        ]
    conn.close()
    # Indicadores sempre sobre o território inteiro — o filtro recorta apenas
    # a lista de casas.
    stats = build_dashboard_stats(casas, pacientes)

    # Filtro de casas: por tipo de imóvel (multi) e/ou por quadra.
    tipos_filtro = [
        tipo for tipo in request.args.getlist("tipo") if tipo in TIPOS_IMOVEL_POR_CODIGO
    ]
    quadra_filtro = request.args.get("quadra", "").strip()

    # Contagens exibidas no modal — calculadas ANTES do recorte, para o
    # operador saber quantos imóveis existem de cada tipo/quadra.
    contagem_tipos = {opcao["codigo"]: 0 for opcao in TIPOS_IMOVEL_OPCOES}
    sem_quadra_total = 0
    for casa in casas:
        contagem_tipos[tipo_imovel_de(casa)] += 1
        if casa["quadra_id"] is None:
            sem_quadra_total += 1

    total_casas_geral = len(casas)
    casas_filtradas = casas
    if tipos_filtro:
        tipos_ativos = set(tipos_filtro)
        casas_filtradas = [casa for casa in casas_filtradas if tipo_imovel_de(casa) in tipos_ativos]
    if quadra_filtro == "0":
        casas_filtradas = [casa for casa in casas_filtradas if casa["quadra_id"] is None]
    elif quadra_filtro.isdigit():
        casas_filtradas = [casa for casa in casas_filtradas if casa["quadra_id"] == int(quadra_filtro)]
    else:
        quadra_filtro = ""

    return render_template(
        "index.html",
        casas=casas_filtradas,
        quadras=quadras,
        stats=stats,
        busca=busca,
        ordem=ordem,
        pacientes_busca=pacientes_busca,
        tipos_filtro=tipos_filtro,
        quadra_filtro=quadra_filtro,
        contagem_tipos=contagem_tipos,
        sem_quadra_total=sem_quadra_total,
        total_casas_geral=total_casas_geral,
        filtro_ativo=bool(tipos_filtro or quadra_filtro),
    )


@app.route("/quadra/nova", methods=["GET", "POST"])
@login_required
def cadastrar_quadra():
    if request.method == "POST":
        numero_quadra, error = parse_positive_int(request.form.get("numero_quadra"), "o número da quadra")

        if error:
            flash(error, "danger")
            return render_template("cadastrar_quadra.html", proximo_numero=get_next_block_number())

        conn = get_db_connection()
        conn.execute("INSERT INTO quadras (numero_quadra) VALUES (?)", (numero_quadra,))
        conn.commit()
        conn.close()
        flash("Quadra cadastrada com sucesso.", "success")
        return redirect(url_for("index"))

    return render_template("cadastrar_quadra.html", proximo_numero=get_next_block_number())


@app.route("/quadra/<int:quadra_id>/editar", methods=["GET", "POST"])
@login_required
def editar_quadra(quadra_id):
    conn = get_db_connection()
    quadra = conn.execute("SELECT * FROM quadras WHERE id = ?", (quadra_id,)).fetchone()

    if quadra is None:
        conn.close()
        flash("Quadra não encontrada.", "warning")
        return redirect(url_for("index"))

    if request.method == "POST":
        numero_quadra, error = parse_positive_int(request.form.get("numero_quadra"), "o número da quadra")
        if error:
            conn.close()
            flash(error, "danger")
            return render_template("editar_quadra.html", quadra=quadra)

        conn.execute("UPDATE quadras SET numero_quadra = ? WHERE id = ?", (numero_quadra, quadra_id))
        conn.commit()
        conn.close()
        flash("Quadra atualizada com sucesso.", "success")
        return redirect(url_for("index"))

    conn.close()
    return render_template("editar_quadra.html", quadra=quadra)


@app.route("/quadra/<int:quadra_id>/excluir", methods=["POST"])
@login_required
def excluir_quadra(quadra_id):
    try:
        criar_backup("antes_excluir_quadra")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de excluir quadra %s: %s", quadra_id, exc)
        flash("Não foi possível criar backup de segurança. Exclusão cancelada.", "danger")
        return redirect(url_for("index"))

    conn = get_db_connection()
    conn.execute("UPDATE casas SET quadra_id = NULL WHERE quadra_id = ?", (quadra_id,))
    conn.execute("DELETE FROM quadras WHERE id = ?", (quadra_id,))
    conn.commit()
    conn.close()
    flash("Quadra excluída. As casas vinculadas ficaram sem quadra.", "success")
    return redirect(url_for("index"))


@app.route("/casa/nova", methods=["GET", "POST"])
@login_required
def cadastrar_casa():
    quadras = get_quadras()
    if request.method == "POST":
        endereco = request.form.get("endereco", "").strip()
        numero_casa_text = request.form.get("numero_casa", "").strip()
        tipo_imovel = normalizar_tipo_imovel(request.form.get("tipo_imovel"))
        quadra_id, quadra_error = parse_positive_int(request.form.get("quadra_id"), "a quadra", required=False)

        if not endereco:
            flash("Informe o endereço da casa.", "danger")
            return render_template(
                "cadastrar_casa.html",
                proximo_numero=get_next_house_number(),
                endereco=endereco,
                numero_casa=numero_casa_text,
                quadra_id=quadra_id,
                quadras=quadras,
            )

        if quadra_error:
            flash(quadra_error, "danger")
            return render_template(
                "cadastrar_casa.html",
                proximo_numero=get_next_house_number(),
                endereco=endereco,
                numero_casa=numero_casa_text,
                quadra_id=None,
                quadras=quadras,
            )

        if not quadra_exists(quadra_id):
            flash("Quadra selecionada não existe.", "danger")
            return render_template(
                "cadastrar_casa.html",
                proximo_numero=get_next_house_number(),
                endereco=endereco,
                numero_casa=numero_casa_text,
                quadra_id=None,
                quadras=quadras,
            )

        numero_final, numero_error = parse_positive_int(numero_casa_text, "o número da casa", required=False)
        if numero_error:
            flash(numero_error, "danger")
            return render_template(
                "cadastrar_casa.html",
                proximo_numero=get_next_house_number(quadra_id),
                endereco=endereco,
                numero_casa=numero_casa_text,
                quadra_id=quadra_id,
                quadras=quadras,
            )

        if numero_final is None:
            numero_final = get_next_house_number(quadra_id)

        conn = get_db_connection()
        conn.execute(
            "INSERT INTO casas (quadra_id, numero_casa, endereco, tipo_imovel) VALUES (?, ?, ?, ?)",
            (quadra_id, numero_final, endereco, tipo_imovel),
        )
        conn.commit()
        conn.close()
        flash("Casa cadastrada com sucesso.", "success")
        return redirect(url_for("index"))

    return render_template("cadastrar_casa.html", proximo_numero=get_next_house_number(), quadras=quadras)


@app.route("/casa/<int:casa_id>")
@login_required
def detalhes_casa(casa_id):
    casa = get_house_or_404(casa_id)
    if casa is None:
        return redirect(url_for("index"))

    conn = get_db_connection()
    moradores = conn.execute(
        "SELECT * FROM pacientes WHERE casa_id = ? ORDER BY nome",
        (casa_id,),
    ).fetchall()
    conn.close()
    pacientes = [p for p in moradores if status_paciente_de(p) == "ativo"]
    pacientes_inativos = [p for p in moradores if status_paciente_de(p) != "ativo"]
    return render_template(
        "detalhes_casa.html",
        casa=casa,
        pacientes=pacientes,
        pacientes_inativos=pacientes_inativos,
        casas_destino=get_casas_para_transferencia(excluir_casa_id=casa_id),
    )


@app.route("/casa/<int:casa_id>/editar", methods=["GET", "POST"])
@login_required
def editar_casa(casa_id):
    casa = get_house_or_404(casa_id)
    if casa is None:
        return redirect(url_for("index"))

    quadras = get_quadras()
    if request.method == "POST":
        numero_casa_text = request.form.get("numero_casa", "").strip()
        endereco = request.form.get("endereco", "").strip()
        tipo_imovel = normalizar_tipo_imovel(request.form.get("tipo_imovel"))
        quadra_id, quadra_error = parse_positive_int(request.form.get("quadra_id"), "a quadra", required=False)

        if not endereco:
            flash("Informe o endereço da casa.", "danger")
            return render_template("editar_casa.html", casa=casa, quadras=quadras)

        if quadra_error:
            flash(quadra_error, "danger")
            return render_template("editar_casa.html", casa=casa, quadras=quadras)

        if not quadra_exists(quadra_id):
            flash("Quadra selecionada não existe.", "danger")
            return render_template("editar_casa.html", casa=casa, quadras=quadras)

        numero_casa, numero_error = parse_positive_int(numero_casa_text, "o número da casa")
        if numero_error:
            flash(numero_error, "danger")
            return render_template("editar_casa.html", casa=casa, quadras=quadras)

        conn = get_db_connection()
        conn.execute(
            "UPDATE casas SET quadra_id = ?, numero_casa = ?, endereco = ?, tipo_imovel = ? WHERE id = ?",
            (quadra_id, numero_casa, endereco, tipo_imovel, casa_id),
        )
        conn.commit()
        conn.close()
        flash("Casa atualizada com sucesso.", "success")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    return render_template("editar_casa.html", casa=casa, quadras=quadras)


@app.route("/casa/<int:casa_id>/excluir", methods=["POST"])
@login_required
def excluir_casa(casa_id):
    try:
        criar_backup("antes_excluir_casa")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de excluir casa %s: %s", casa_id, exc)
        flash("Não foi possível criar backup de segurança. Exclusão cancelada.", "danger")
        return redirect(url_for("index"))

    conn = get_db_connection()
    conn.execute("DELETE FROM casas WHERE id = ?", (casa_id,))
    conn.commit()
    conn.close()
    flash("Casa e pacientes vinculados foram excluídos.", "success")
    return redirect(url_for("index"))


@app.route("/casa/<int:casa_id>/paciente/novo", methods=["GET", "POST"])
@login_required
def cadastrar_paciente(casa_id):
    casa = get_house_or_404(casa_id)
    if casa is None:
        return redirect(url_for("index"))

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        if not nome:
            flash("Informe o nome do paciente.", "danger")
            return render_template("cadastrar_paciente.html", casa=casa, paciente=request.form)

        duplicado = paciente_com_documento(apenas_digitos(request.form.get("cpf", "")))
        if duplicado:
            flash(
                f"Este CPF/CNS já está cadastrado para {duplicado['nome']}. "
                "Cada documento identifica um único paciente — localize o cadastro na página Pacientes.",
                "danger",
            )
            return render_template("cadastrar_paciente.html", casa=casa, paciente=request.form)

        inserir_paciente_do_form(casa_id, nome)
        flash("Paciente cadastrado com sucesso.", "success")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    return render_template("cadastrar_paciente.html", casa=casa)


@app.route("/pacientes/novo", methods=["GET", "POST"])
@login_required
def cadastrar_paciente_geral():
    """Cadastro a partir da lista de pacientes: a casa é opcional — dá para
    cadastrar sem casa e vincular depois pelo "Definir casa"."""
    casas_opcoes = get_casas_para_transferencia()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        casa_id_raw = request.form.get("casa_id", "").strip()

        def render_erro():
            return render_template(
                "cadastrar_paciente_geral.html",
                casa=None,
                paciente=request.form,
                casas_opcoes=casas_opcoes,
                casa_id_selecionada=casa_id_raw,
            )

        if not nome:
            flash("Informe o nome do paciente.", "danger")
            return render_erro()

        casa_id = None
        if casa_id_raw:
            casa_id, casa_error = parse_positive_int(casa_id_raw, "a casa", required=False)
            if casa_error:
                flash(casa_error, "danger")
                return render_erro()
            conn = get_db_connection()
            existe = conn.execute("SELECT 1 FROM casas WHERE id = ?", (casa_id,)).fetchone()
            conn.close()
            if existe is None:
                flash("Casa selecionada não existe.", "danger")
                return render_erro()

        duplicado = paciente_com_documento(apenas_digitos(request.form.get("cpf", "")))
        if duplicado:
            flash(
                f"Este CPF/CNS já está cadastrado para {duplicado['nome']}. "
                "Cada documento identifica um único paciente.",
                "danger",
            )
            return render_erro()

        inserir_paciente_do_form(casa_id, nome)
        flash("Paciente cadastrado com sucesso.", "success")
        return redirect(url_for("listar_pacientes"))

    return render_template(
        "cadastrar_paciente_geral.html",
        casa=None,
        casas_opcoes=casas_opcoes,
        casa_id_selecionada="",
    )


@app.route("/paciente/<int:paciente_id>/editar", methods=["GET", "POST"])
@login_required
def editar_paciente(paciente_id):
    conn = get_db_connection()
    paciente = conn.execute("SELECT * FROM pacientes WHERE id = ?", (paciente_id,)).fetchone()

    if paciente is None:
        conn.close()
        flash("Paciente não encontrado.", "warning")
        return redirect(url_for("index"))

    casa = conn.execute("SELECT * FROM casas WHERE id = ?", (paciente["casa_id"],)).fetchone()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        if not nome:
            flash("Informe o nome do paciente.", "danger")
            conn.close()
            return render_template("editar_paciente.html", paciente=paciente, casa=casa)

        duplicado = paciente_com_documento(
            apenas_digitos(request.form.get("cpf", "")), excluir_id=paciente_id
        )
        if duplicado:
            flash(
                f"Este CPF/CNS já está cadastrado para {duplicado['nome']}. "
                "Cada documento identifica um único paciente.",
                "danger",
            )
            conn.close()
            return render_template("editar_paciente.html", paciente=paciente, casa=casa)

        conn.execute(
            """
            UPDATE pacientes
            SET nome = ?, cpf = ?, telefone = ?, data_nascimento = ?, nome_pai = ?,
                nome_mae = ?, sexo = ?, condicoes_saude = ?, observacao = ?
            WHERE id = ?
            """,
            (
                nome,
                formatar_cpf_ou_cns(request.form.get("cpf", "")),
                formatar_telefone(request.form.get("telefone", "")),
                request.form.get("data_nascimento", "").strip(),
                request.form.get("nome_pai", "").strip(),
                request.form.get("nome_mae", "").strip(),
                request.form.get("sexo", "").strip(),
                normalizar_condicoes(request.form.getlist("condicoes_saude")),
                request.form.get("observacao", "").strip(),
                paciente_id,
            ),
        )
        conn.commit()
        casa_id = paciente["casa_id"]
        conn.close()
        flash("Paciente atualizado com sucesso.", "success")
        return redirect_apos_acao_no_paciente(casa_id)

    conn.close()
    return render_template("editar_paciente.html", paciente=paciente, casa=casa)


# ---------------------------------------------------------------------------
# Lixeira de pacientes
#
# Excluir um paciente é mover para cá — restaurável por 30 dias. Depois disso
# (ou ao esvaziar a lixeira, com backup antes), o registro sai de vez.
# ---------------------------------------------------------------------------
LIXEIRA_RETENCAO_DIAS = 30

_COLUNAS_PACIENTE = (
    "casa_id, nome, cpf, telefone, data_nascimento, sexo, "
    "nome_pai, nome_mae, condicoes_saude, observacao, status"
)


def purgar_lixeira_expirada(conn):
    """Remove da lixeira o que passou da retenção. Chamado no boot e ao abrir
    a página — o prazo de 30 dias é honrado sem tarefa agendada."""
    limite = (datetime.now() - timedelta(days=LIXEIRA_RETENCAO_DIAS)).isoformat(timespec="seconds")
    cursor = conn.execute("DELETE FROM lixeira_pacientes WHERE excluido_em < ?", (limite,))
    if cursor.rowcount:
        logger.info("Lixeira: %s registro(s) expirados purgados.", cursor.rowcount)
    return cursor.rowcount


@app.route("/paciente/<int:paciente_id>/excluir", methods=["POST"])
@login_required
def excluir_paciente(paciente_id):
    conn = get_db_connection()
    paciente = conn.execute("SELECT * FROM pacientes WHERE id = ?", (paciente_id,)).fetchone()

    if paciente is None:
        conn.close()
        flash("Paciente não encontrado.", "warning")
        return redirect(url_for("index"))

    casa_id = paciente["casa_id"]
    # Mover para a lixeira e remover da tabela ativa — mesma transação.
    conn.execute(
        f"""
        INSERT INTO lixeira_pacientes (id, {_COLUNAS_PACIENTE}, excluido_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            paciente["id"],
            paciente["casa_id"],
            paciente["nome"],
            paciente["cpf"],
            paciente["telefone"],
            paciente["data_nascimento"],
            paciente["sexo"],
            paciente["nome_pai"],
            paciente["nome_mae"],
            paciente["condicoes_saude"],
            paciente["observacao"],
            status_paciente_de(paciente),
            datetime.now().isoformat(timespec="seconds"),
        ),
    )
    conn.execute("DELETE FROM pacientes WHERE id = ?", (paciente_id,))
    conn.commit()
    conn.close()
    flash(
        f"{paciente['nome']} foi para a Lixeira — restaurável por {LIXEIRA_RETENCAO_DIAS} dias.",
        "success",
    )
    # Exclusão disparada da lista de pacientes volta para a própria lista
    # (com filtros/página preservados); das demais telas, para a casa.
    if request.form.get("next"):
        return redirect(proxima_url_segura(request.form.get("next")))
    return redirect_apos_acao_no_paciente(casa_id)


@app.route("/lixeira")
@login_required
def lixeira():
    conn = get_db_connection()
    purgar_lixeira_expirada(conn)
    conn.commit()
    registros = conn.execute(
        "SELECT * FROM lixeira_pacientes ORDER BY excluido_em DESC, id DESC"
    ).fetchall()
    conn.close()

    agora = datetime.now()
    itens = []
    for registro in registros:
        try:
            excluido_em = datetime.fromisoformat(registro["excluido_em"])
        except ValueError:
            excluido_em = agora
        dias_restantes = max(0, LIXEIRA_RETENCAO_DIAS - (agora - excluido_em).days)
        itens.append({**dict(registro), "dias_restantes": dias_restantes, "excluido_em_dt": excluido_em})

    return render_template("lixeira.html", itens=itens, retencao=LIXEIRA_RETENCAO_DIAS)


@app.route("/lixeira/<int:paciente_id>/restaurar", methods=["POST"])
@login_required
def restaurar_da_lixeira(paciente_id):
    conn = get_db_connection()
    item = conn.execute(
        "SELECT * FROM lixeira_pacientes WHERE id = ?", (paciente_id,)
    ).fetchone()

    if item is None:
        conn.close()
        flash("Registro não encontrado na lixeira.", "warning")
        return redirect(url_for("lixeira"))

    # A mesma regra de unicidade do cadastro: se o CPF/CNS foi recadastrado
    # enquanto o registro esperava na lixeira, não pode haver dois.
    conn.close()
    duplicado = paciente_com_documento(apenas_digitos(item["cpf"]))
    if duplicado:
        flash(
            f"Este CPF/CNS agora pertence ao cadastro de {duplicado['nome']}. "
            "Resolva o cadastro atual antes de restaurar.",
            "danger",
        )
        return redirect(url_for("lixeira"))

    conn = get_db_connection()
    casa_id = item["casa_id"]
    aviso_casa = ""
    if casa_id is not None:
        existe = conn.execute("SELECT 1 FROM casas WHERE id = ?", (casa_id,)).fetchone()
        if existe is None:
            casa_id = None
            aviso_casa = " A casa original não existe mais — o paciente voltou sem casa."

    valores = (
        casa_id,
        item["nome"],
        item["cpf"],
        item["telefone"],
        item["data_nascimento"],
        item["sexo"],
        item["nome_pai"],
        item["nome_mae"],
        item["condicoes_saude"],
        item["observacao"],
        normalizar_status_paciente(item["status"]),
    )
    try:
        # Preserva o id original quando livre (AUTOINCREMENT nunca o reusa).
        conn.execute(
            f"INSERT INTO pacientes (id, {_COLUNAS_PACIENTE}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (item["id"], *valores),
        )
    except sqlite3.IntegrityError:
        conn.execute(
            f"INSERT INTO pacientes ({_COLUNAS_PACIENTE}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            valores,
        )
    conn.execute("DELETE FROM lixeira_pacientes WHERE id = ?", (paciente_id,))
    conn.commit()
    conn.close()
    flash(f"{item['nome']} restaurado com sucesso.{aviso_casa}", "success")
    return redirect(url_for("lixeira"))


@app.route("/lixeira/esvaziar", methods=["POST"])
@login_required
def esvaziar_lixeira():
    try:
        criar_backup("antes_esvaziar_lixeira")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de esvaziar a lixeira: %s", exc)
        flash("Não foi possível criar backup de segurança. A lixeira não foi esvaziada.", "danger")
        return redirect(url_for("lixeira"))

    conn = get_db_connection()
    cursor = conn.execute("DELETE FROM lixeira_pacientes")
    conn.commit()
    conn.close()

    if cursor.rowcount == 0:
        flash("A lixeira já estava vazia.", "info")
    else:
        logger.info("Lixeira esvaziada (%s registros) a partir de %s", cursor.rowcount, request.remote_addr)
        flash(f"Lixeira esvaziada — {cursor.rowcount} registro(s) removidos em definitivo.", "success")
    return redirect(url_for("lixeira"))


# Paginação da lista de pacientes: tamanhos permitidos e padrão.
POR_PAGINA_OPCOES = (25, 50, 100, 200)
POR_PAGINA_PADRAO = 50


def _janela_paginacao(pagina, total_paginas, vizinhos=1):
    """Números de página exibidos: primeira, última e a vizinhança da atual,
    com None marcando as elipses. Até 7 páginas, mostra todas."""
    if total_paginas <= 7:
        return list(range(1, total_paginas + 1))

    visiveis = {1, total_paginas}
    visiveis.update(
        p for p in range(pagina - vizinhos, pagina + vizinhos + 1) if 1 <= p <= total_paginas
    )
    janela = []
    anterior = None
    for numero in sorted(visiveis):
        if anterior is not None and numero - anterior > 1:
            janela.append(None)
        janela.append(numero)
        anterior = numero
    return janela


@app.route("/pacientes")
@login_required
def listar_pacientes():
    busca = request.args.get("busca", "").strip()
    status_filtro = request.args.get("status", "").strip()
    if status_filtro not in STATUS_PACIENTE_POR_CODIGO:
        status_filtro = ""
    quadra_filtro = request.args.get("quadra", "").strip()

    conn = get_db_connection()
    pacientes = conn.execute(
        """
        SELECT pacientes.*, casas.numero_casa, casas.endereco, casas.quadra_id, quadras.numero_quadra
        FROM pacientes
        LEFT JOIN casas ON casas.id = pacientes.casa_id
        LEFT JOIN quadras ON quadras.id = casas.quadra_id
        ORDER BY pacientes.nome COLLATE NOCASE, pacientes.id
        """
    ).fetchall()
    quadras = conn.execute("SELECT * FROM quadras ORDER BY numero_quadra, id").fetchall()
    conn.close()

    # Contagens sempre sobre o cadastro inteiro — o filtro recorta só a lista.
    total_geral = len(pacientes)
    contagem_status = {opcao["codigo"]: 0 for opcao in STATUS_PACIENTE_OPCOES}
    for paciente in pacientes:
        contagem_status[status_paciente_de(paciente)] += 1

    filtrados = pacientes
    if status_filtro:
        filtrados = [p for p in filtrados if status_paciente_de(p) == status_filtro]
    if quadra_filtro == "0":
        filtrados = [p for p in filtrados if p["quadra_id"] is None]
    elif quadra_filtro.isdigit():
        filtrados = [p for p in filtrados if p["quadra_id"] == int(quadra_filtro)]
    else:
        quadra_filtro = ""
    if busca:
        filtrados = [p for p in filtrados if paciente_corresponde_busca(p, busca)]

    # Paginação server-side, aplicada DEPOIS dos filtros. Valores fora da
    # whitelist/intervalo caem no padrão — nunca 500 por query manipulada.
    # "todos" desliga a paginação e mostra a lista inteira.
    por_pagina_raw = request.args.get("por_pagina", "")
    ver_todos = por_pagina_raw == "todos"
    por_pagina = (
        int(por_pagina_raw)
        if por_pagina_raw.isdigit() and int(por_pagina_raw) in POR_PAGINA_OPCOES
        else POR_PAGINA_PADRAO
    )
    pagina_raw = request.args.get("pagina", "")
    pagina = int(pagina_raw) if pagina_raw.isdigit() and int(pagina_raw) >= 1 else 1

    total_filtrado = len(filtrados)
    if ver_todos:
        pagina = 1
        total_paginas = 1
        inicio = 0
        pagina_de_pacientes = filtrados
    else:
        total_paginas = max(1, -(-total_filtrado // por_pagina))
        pagina = min(pagina, total_paginas)
        inicio = (pagina - 1) * por_pagina
        pagina_de_pacientes = filtrados[inicio:inicio + por_pagina]

    return render_template(
        "pacientes.html",
        pacientes=pagina_de_pacientes,
        total_geral=total_geral,
        total_filtrado=total_filtrado,
        contagem_status=contagem_status,
        quadras=quadras,
        casas_destino=get_casas_para_transferencia(),
        busca=busca,
        status_filtro=status_filtro,
        quadra_filtro=quadra_filtro,
        filtro_ativo=bool(busca or status_filtro or quadra_filtro),
        pagina=pagina,
        total_paginas=total_paginas,
        por_pagina_token="todos" if ver_todos else str(por_pagina),
        por_pagina_padrao=str(POR_PAGINA_PADRAO),
        opcoes_por_pagina=POR_PAGINA_OPCOES,
        paginas_visiveis=_janela_paginacao(pagina, total_paginas),
        inicio_exibicao=inicio + 1 if pagina_de_pacientes else 0,
        fim_exibicao=inicio + len(pagina_de_pacientes),
    )


@app.route("/paciente/<int:paciente_id>/status", methods=["POST"])
@login_required
def alterar_status_paciente(paciente_id):
    destino = proxima_url_segura(request.form.get("next"))
    status = request.form.get("status", "").strip()
    if status not in STATUS_PACIENTE_POR_CODIGO:
        flash("Situação inválida.", "danger")
        return redirect(destino)

    conn = get_db_connection()
    paciente = conn.execute("SELECT nome FROM pacientes WHERE id = ?", (paciente_id,)).fetchone()
    if paciente is None:
        conn.close()
        flash("Paciente não encontrado.", "warning")
        return redirect(destino)

    conn.execute("UPDATE pacientes SET status = ? WHERE id = ?", (status, paciente_id))
    conn.commit()
    conn.close()
    if status == "ativo":
        flash(f"{paciente['nome']} voltou a contar como morador ativo.", "success")
    else:
        flash(
            f"{paciente['nome']} marcado como “{status_paciente_label(status)}”. "
            "O cadastro fica guardado, fora das contagens do território.",
            "success",
        )
    return redirect(destino)


@app.route("/pacientes/status-em-massa", methods=["POST"])
@login_required
def alterar_status_em_massa():
    destino = proxima_url_segura(request.form.get("next"))
    status = request.form.get("status", "").strip()
    if status not in STATUS_PACIENTE_POR_CODIGO:
        flash("Escolha a situação a aplicar.", "danger")
        return redirect(destino)

    ids = []
    for valor in request.form.getlist("paciente_ids"):
        try:
            ids.append(int(valor))
        except ValueError:
            continue
    if not ids:
        flash("Selecione pelo menos um paciente.", "warning")
        return redirect(destino)

    # Mutação em lote: backup antes, como toda operação de largo alcance.
    try:
        criar_backup("antes_status_em_massa")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de alterar status em massa: %s", exc)
        flash("Não foi possível criar backup de segurança. Alteração cancelada.", "danger")
        return redirect(destino)

    conn = get_db_connection()
    alterados = 0
    # Em blocos: o limite de parâmetros do SQLite nunca é atingido, qualquer
    # que seja o tamanho da seleção.
    for inicio in range(0, len(ids), 500):
        bloco = ids[inicio:inicio + 500]
        marcadores = ",".join("?" * len(bloco))
        cursor = conn.execute(
            f"UPDATE pacientes SET status = ? WHERE id IN ({marcadores})",
            [status, *bloco],
        )
        alterados += cursor.rowcount
    conn.commit()
    conn.close()

    logger.info(
        "Status em massa (%s): %s paciente(s) → %s", request.remote_addr, alterados, status
    )
    flash(
        f"{alterados} paciente(s) agora como “{status_paciente_label(status)}”. "
        "Cadastros nunca são apagados — dá para reverter um a um ou em massa.",
        "success",
    )
    return redirect(destino)


@app.route("/paciente/<int:paciente_id>/transferir", methods=["POST"])
@login_required
def transferir_paciente(paciente_id):
    destino_url = proxima_url_segura(request.form.get("next"))
    casa_destino_id, error = parse_positive_int(request.form.get("casa_destino_id"), "a casa de destino")
    if error:
        flash(error, "danger")
        return redirect(destino_url)

    conn = get_db_connection()
    paciente = conn.execute(
        "SELECT nome, casa_id, status FROM pacientes WHERE id = ?", (paciente_id,)
    ).fetchone()
    if paciente is None:
        conn.close()
        flash("Paciente não encontrado.", "warning")
        return redirect(destino_url)

    if status_paciente_de(paciente) == "obito":
        conn.close()
        flash(
            "Paciente com óbito registrado não pode ser transferido. "
            "Se for correção de cadastro, reative o paciente primeiro.",
            "warning",
        )
        return redirect(destino_url)

    casa_destino = conn.execute(
        "SELECT id, numero_casa FROM casas WHERE id = ?", (casa_destino_id,)
    ).fetchone()
    if casa_destino is None:
        conn.close()
        flash("Casa de destino não encontrada.", "warning")
        return redirect(destino_url)

    if paciente["casa_id"] == casa_destino_id:
        conn.close()
        flash("O paciente já está nessa casa.", "info")
        return redirect(destino_url)

    # Transferir dentro da área = novo endereço conhecido → volta a ser ativo.
    # A condição repete o bloqueio de óbito no próprio UPDATE: mesmo que o
    # status mude entre o SELECT acima e aqui (outra aba), nada é "ressuscitado".
    cursor = conn.execute(
        "UPDATE pacientes SET casa_id = ?, status = 'ativo' "
        "WHERE id = ? AND COALESCE(status, 'ativo') != 'obito'",
        (casa_destino_id, paciente_id),
    )
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        flash("A transferência não foi aplicada — o cadastro mudou nesse meio tempo.", "warning")
        return redirect(destino_url)
    flash(
        f"{paciente['nome']} transferido para a casa {casa_destino['numero_casa']}.",
        "success",
    )
    return redirect(destino_url)


@app.route("/casa/<int:casa_id>/transferir", methods=["POST"])
@login_required
def transferir_familia(casa_id):
    casa = get_house_or_404(casa_id)
    if casa is None:
        return redirect(url_for("index"))

    casa_destino_id, error = parse_positive_int(request.form.get("casa_destino_id"), "a casa de destino")
    if error:
        flash(error, "danger")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    if casa_destino_id == casa_id:
        flash("Escolha uma casa diferente da atual.", "danger")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    conn = get_db_connection()
    casa_destino = conn.execute(
        "SELECT id, numero_casa FROM casas WHERE id = ?", (casa_destino_id,)
    ).fetchone()
    conn.close()
    if casa_destino is None:
        flash("Casa de destino não encontrada.", "warning")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    try:
        criar_backup("antes_transferir_familia")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de transferir família da casa %s: %s", casa_id, exc)
        flash("Não foi possível criar backup de segurança. Transferência cancelada.", "danger")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    # Só moradores ativos mudam junto — quem já se mudou antes fica registrado
    # na casa onde morava.
    conn = get_db_connection()
    cursor = conn.execute(
        f"UPDATE pacientes SET casa_id = ? WHERE casa_id = ? AND {SQL_PACIENTE_ATIVO}",
        (casa_destino_id, casa_id),
    )
    transferidos = cursor.rowcount
    conn.commit()
    conn.close()

    if transferidos == 0:
        flash("Esta casa não tem morador ativo para transferir.", "info")
        return redirect(url_for("detalhes_casa", casa_id=casa_id))

    flash(
        f"{transferidos} morador(es) transferido(s) para a casa {casa_destino['numero_casa']}.",
        "success",
    )
    return redirect(url_for("detalhes_casa", casa_id=casa_destino_id))


# ---------------------------------------------------------------------------
# Importação de pacientes a partir do CSV do e-SUS APS
#
# O relatório "acompanhamento de condições de saúde" vem com preâmbulo de
# filtros, separador ';' e codificação Windows (cp1252). O parser localiza o
# cabeçalho pelo nome das colunas — não por posição fixa — para sobreviver a
# variações entre versões do e-SUS.
# ---------------------------------------------------------------------------
class ImportacaoInvalida(ValueError):
    """Arquivo enviado não é um relatório CSV utilizável do e-SUS."""


def _csv_valor(value):
    value = str(value or "").strip()
    return "" if value in ("-", "—") else value


def _decodificar_csv(dados):
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return dados.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportacaoInvalida("Não foi possível ler o arquivo — codificação de texto desconhecida.")


def _data_iso(valor):
    try:
        return datetime.strptime(valor, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def parse_relatorio_esus(texto):
    """Extrai os pacientes do CSV do e-SUS: localiza a linha de cabeçalho,
    mapeia as colunas por nome (sem depender de acento/caixa) e devolve
    registros prontos para inserção."""
    linhas = list(csv.reader(texto.splitlines(), delimiter=";"))

    cabecalho_idx = None
    for idx, linha in enumerate(linhas):
        if linha and texto_normalizado(linha[0]) == "nome" and len(linha) > 3:
            cabecalho_idx = idx
            break
    if cabecalho_idx is None:
        raise ImportacaoInvalida(
            "Cabeçalho não encontrado. Envie o CSV exportado do e-SUS "
            "(relatório de acompanhamento de condições de saúde)."
        )

    colunas = {texto_normalizado(nome): pos for pos, nome in enumerate(linhas[cabecalho_idx])}

    def campo(linha, nome):
        pos = colunas.get(nome)
        if pos is None or pos >= len(linha):
            return ""
        return _csv_valor(linha[pos])

    registros = []
    for linha in linhas[cabecalho_idx + 1:]:
        if not linha or len(linha) < 4:
            continue
        nome = campo(linha, "nome")
        if not nome:
            continue

        cpf = apenas_digitos(campo(linha, "cpf"))
        cns = apenas_digitos(campo(linha, "cns"))
        documento = cpf or cns
        data_nascimento = _data_iso(campo(linha, "data de nascimento"))

        # Linha sem documento e sem nascimento não é um cadastro do e-SUS —
        # protege contra rodapés/totalizadores de outras versões do relatório.
        if not documento and not data_nascimento:
            continue

        telefone = (
            campo(linha, "telefone celular")
            or campo(linha, "telefone residencial")
            or campo(linha, "telefone de contato")
        )

        sexo = campo(linha, "sexo").capitalize()
        if sexo not in ("Masculino", "Feminino"):
            sexo = ""

        registros.append(
            {
                "nome": nome,
                "cpf": formatar_cpf_ou_cns(documento),
                "documento_digitos": documento,
                "telefone": formatar_telefone(telefone),
                "data_nascimento": data_nascimento,
                "sexo": sexo,
            }
        )

    return registros


def _inserir_pacientes_importados(registros):
    """Insere os registros SEM vincular a casa nenhuma — o vínculo é decisão
    do operador, feita depois pelo "Definir casa" na lista de pacientes.
    Nenhuma observação é gerada: o cadastro entra limpo.
    Deduplicação: CPF/CNS já cadastrado, ou mesmo nome + data de nascimento."""
    conn = get_db_connection()
    try:
        existentes = conn.execute("SELECT nome, cpf, data_nascimento FROM pacientes").fetchall()
        docs_existentes = {apenas_digitos(p["cpf"]) for p in existentes if apenas_digitos(p["cpf"])}
        nomes_existentes = {
            (texto_normalizado(p["nome"]), p["data_nascimento"] or "") for p in existentes
        }

        resultado = {"importados": 0, "ignorados": 0}

        for registro in registros:
            documento = registro["documento_digitos"]
            chave_nome = (texto_normalizado(registro["nome"]), registro["data_nascimento"] or "")
            if (documento and documento in docs_existentes) or chave_nome in nomes_existentes:
                resultado["ignorados"] += 1
                continue

            conn.execute(
                """
                INSERT INTO pacientes (casa_id, nome, cpf, telefone, data_nascimento, sexo,
                                       nome_pai, nome_mae, condicoes_saude, observacao)
                VALUES (NULL, ?, ?, ?, ?, ?, '', '', '', '')
                """,
                (
                    registro["nome"],
                    registro["cpf"],
                    registro["telefone"],
                    registro["data_nascimento"],
                    registro["sexo"],
                ),
            )
            if documento:
                docs_existentes.add(documento)
            nomes_existentes.add(chave_nome)
            resultado["importados"] += 1

        conn.commit()
        return resultado
    finally:
        conn.close()


@app.route("/pacientes/importar", methods=["GET", "POST"])
@login_required
def importar_pacientes():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if arquivo is None or not arquivo.filename:
            flash("Selecione o arquivo CSV exportado do e-SUS.", "danger")
            return redirect(url_for("importar_pacientes"))

        try:
            registros = parse_relatorio_esus(_decodificar_csv(arquivo.read()))
        except ImportacaoInvalida as exc:
            logger.warning("Importação e-SUS rejeitada (%s): %s", request.remote_addr, exc)
            flash(str(exc), "danger")
            return redirect(url_for("importar_pacientes"))

        if not registros:
            flash("Nenhum paciente encontrado no arquivo.", "warning")
            return redirect(url_for("importar_pacientes"))

        try:
            criar_backup("antes_importar_pacientes")
        except (sqlite3.Error, OSError) as exc:
            logger.error("Falha ao criar backup antes de importar pacientes: %s", exc)
            flash("Não foi possível criar backup de segurança. Importação cancelada.", "danger")
            return redirect(url_for("importar_pacientes"))

        try:
            resultado = _inserir_pacientes_importados(registros)
        except sqlite3.Error as exc:
            logger.error("Falha ao importar pacientes do e-SUS: %s", exc)
            flash("Não foi possível concluir a importação. Nada foi alterado.", "danger")
            return redirect(url_for("importar_pacientes"))

        logger.info("Importação e-SUS concluída (%s): %s", request.remote_addr, resultado)
        partes = [f"{resultado['importados']} paciente(s) importado(s) sem casa — use “Definir casa” para vincular"]
        if resultado["ignorados"]:
            partes.append(f"{resultado['ignorados']} já cadastrado(s), ignorado(s)")
        flash("Importação concluída: " + "; ".join(partes) + ".", "success")
        return redirect(url_for("listar_pacientes"))

    return render_template("importar_pacientes.html")


@app.route("/exportar")
@login_required
def exportar():
    return render_template(
        "exportar.html",
        opcoes_grupos=GRUPOS_DEMOGRAFICOS_OPCOES,
        acs_nome=get_preferencia("acs_nome"),
        acs_microarea=get_preferencia("acs_microarea"),
    )


@app.route("/exportar/preview")
@login_required
def preview_exportar_pdf():
    condicoes_selecionadas = ler_codigos_condicoes(request.args.getlist("condicoes"))
    grupos_selecionados = ler_codigos_grupos(request.args.getlist("grupos"))
    casas, pacientes = carregar_dados_relatorio(
        condicoes_selecionadas, grupos_selecionados=grupos_selecionados
    )
    stats = build_dashboard_stats(casas, pacientes)
    comorbidades = dict(stats["comorbidades"])
    condicoes_preview = [
        {
            "codigo": codigo,
            "label": condicao_label(codigo),
            "total": comorbidades.get(condicao_label(codigo), 0),
        }
        for codigo in condicoes_selecionadas
    ]
    # Totais por grupo calculados sobre o recorte já filtrado — a prévia
    # descreve exatamente o documento que sai.
    grupos_preview = [
        {
            "codigo": codigo,
            "label": GRUPOS_POR_CODIGO[codigo],
            "total": sum(1 for paciente in pacientes if GRUPOS_PREDICADOS[codigo](paciente)),
        }
        for codigo in grupos_selecionados
    ]

    return jsonify(
        {
            "modo": "filtrado" if (condicoes_selecionadas or grupos_selecionados) else "geral",
            "condicoes": condicoes_preview,
            "grupos": grupos_preview,
            "stats": {
                "total_casas": stats["total_casas"],
                "casas_vazias": stats["casas_vazias"],
                "total_pacientes": stats["total_pacientes"],
                "criancas": stats["criancas"],
                "idosos": stats["idosos"],
                "gestantes": stats["gestantes"],
                "homens": stats["homens"],
                "mulheres": stats["mulheres"],
            },
        }
    )


@app.route("/exportar/pdf")
@login_required
def exportar_pdf():
    condicoes_selecionadas = ler_codigos_condicoes(request.args.getlist("condicoes"))
    grupos_selecionados = ler_codigos_grupos(request.args.getlist("grupos"))
    filtro_ativo = bool(condicoes_selecionadas or grupos_selecionados)
    modo_filtro = request.args.get("filtrar") == "1" or filtro_ativo
    filtro_sem_selecao = modo_filtro and not filtro_ativo
    filtro_labels = [condicao_label(codigo) for codigo in condicoes_selecionadas]
    casas, pacientes = carregar_dados_relatorio(
        condicoes_selecionadas, filtro_sem_selecao, grupos_selecionados
    )

    pacientes_por_casa = {}
    for paciente in pacientes:
        pacientes_por_casa.setdefault(paciente["casa_id"], []).append(paciente)

    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")

    def adicionar_cabecalho_rodape(canvas, doc):
        canvas.saveState()
        if doc.page > 1:
            canvas.setFont("Helvetica-Bold", 8)
            canvas.setFillColor(colors.HexColor("#0b5ed7"))
            canvas.drawString(1.5 * cm, A4[1] - 1.1 * cm, "SAÚDE SIMPLES")
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#495057"))
            largura_marca = canvas.stringWidth("SAÚDE SIMPLES", "Helvetica-Bold", 8)
            canvas.drawString(
                1.5 * cm + largura_marca + 6, A4[1] - 1.1 * cm, "Relatório do Território"
            )
            canvas.setFillColor(colors.HexColor("#6c757d"))
            canvas.drawRightString(A4[0] - 1.5 * cm, A4[1] - 1.1 * cm, f"Gerado em {data_geracao}")
            canvas.setStrokeColor(colors.HexColor("#dee2e6"))
            canvas.line(1.5 * cm, A4[1] - 1.3 * cm, A4[0] - 1.5 * cm, A4[1] - 1.3 * cm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#6c757d"))
        canvas.drawString(1.5 * cm, 0.9 * cm, "Saúde Simples")
        canvas.drawRightString(A4[0] - 1.5 * cm, 0.9 * cm, f"Página {doc.page}")
        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    # Cabeçalho editorial: overline de marca, título forte à esquerda e
    # metadados à direita — hierarquia de documento, não de banner.
    overline_style = ParagraphStyle(
        "OverlineMarca",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#0b5ed7"),
        spaceAfter=3,
    )
    title_style = ParagraphStyle(
        "TituloRelatorio",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=23,
        textColor=colors.HexColor("#212529"),
        alignment=0,
        spaceAfter=0,
    )
    meta_style = ParagraphStyle(
        "MetaRelatorio",
        parent=styles["BodyText"],
        fontSize=8,
        leading=12,
        textColor=colors.HexColor("#6c757d"),
        alignment=2,
    )
    subtitle_style = ParagraphStyle(
        "SubtituloRelatorio",
        parent=styles["BodyText"],
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor("#495057"),
        alignment=0,
        spaceAfter=2,
    )
    caption_style = ParagraphStyle(
        "LegendaImagem",
        parent=styles["BodyText"],
        fontName="Helvetica-Oblique",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#6c757d"),
        alignment=1,
        spaceBefore=4,
    )
    section_style = ParagraphStyle(
        "SecaoRelatorio",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#0b5ed7"),
        spaceBefore=10,
        spaceAfter=8,
    )
    house_style = ParagraphStyle(
        "CasaTitulo",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=colors.white,
    )
    block_style = ParagraphStyle(
        "QuadraTitulo",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0b5ed7"),
        spaceBefore=8,
        spaceAfter=6,
    )
    normal_style = ParagraphStyle(
        "TextoNormal",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#212529"),
    )
    label_style = ParagraphStyle(
        "Rotulo",
        parent=normal_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#495057"),
    )
    number_style = ParagraphStyle(
        "NumeroResumo",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=17,
        textColor=colors.HexColor("#0b5ed7"),
        alignment=1,
    )
    empty_style = ParagraphStyle(
        "Vazio",
        parent=normal_style,
        textColor=colors.HexColor("#6c757d"),
        alignment=1,
    )

    total_pacientes = len(pacientes)
    stats = build_dashboard_stats(casas, pacientes)
    rotulo_casas = "casa" if len(casas) == 1 else "casas"
    rotulo_pacientes = "paciente" if total_pacientes == 1 else "pacientes"
    header_table = Table(
        [
            [
                [
                    Paragraph("SAÚDE SIMPLES", overline_style),
                    Paragraph("Relatório do Território", title_style),
                ],
                Paragraph(
                    f"Gerado em {data_geracao}<br/>"
                    f"{len(casas)} {rotulo_casas} &#183; {total_pacientes} {rotulo_pacientes}",
                    meta_style,
                ),
            ]
        ],
        colWidths=[11.4 * cm, 6.6 * cm],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (0, 0), "TOP"),
                ("VALIGN", (1, 0), (1, 0), "BOTTOM"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    # Régua com acento: segmento curto na cor da marca + linha-fio no restante.
    header_rule = Table([["", ""]], colWidths=[2.6 * cm, 15.4 * cm], rowHeights=[2])
    header_rule.setStyle(
        TableStyle(
            [
                ("LINEBELOW", (0, 0), (0, 0), 1.6, colors.HexColor("#0b5ed7")),
                ("LINEBELOW", (1, 0), (1, 0), 0.5, colors.HexColor("#dee2e6")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements = [
        header_table,
        Spacer(1, 7),
        header_rule,
        Spacer(1, 10),
    ]
    if grupos_selecionados:
        elements.append(
            Paragraph(
                f"<b>Recorte:</b> {xml_escape(rotulo_grupos(grupos_selecionados))}",
                subtitle_style,
            )
        )
    if condicoes_selecionadas:
        elements.append(
            Paragraph(f"<b>Filtro por comorbidade:</b> {', '.join(filtro_labels)}", subtitle_style)
        )
    elif filtro_sem_selecao:
        elements.append(
            Paragraph("<b>Filtro:</b> nenhum grupo ou comorbidade selecionado", subtitle_style)
        )
    if modo_filtro:
        elements.append(Spacer(1, 8))

    report_image_path = os.path.join(BASE_DIR, "image", "image.jpg")
    if os.path.exists(report_image_path):
        # Imagem tratada como figura: moldura fina, legenda discreta abaixo.
        region_image = pdf_image(report_image_path, 17.4 * cm, 4.8 * cm)
        image_frame = Table(
            [[region_image]], colWidths=[region_image.drawWidth + 0.2 * cm]
        )
        image_frame.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dee2e6")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elements.extend(
            [
                image_frame,
                Paragraph("Área de abrangência da UBS", caption_style),
                Spacer(1, 10),
            ]
        )

    summary_top = [
        [
            pdf_text("Total de casas", label_style),
            pdf_text("Casas vazias", label_style),
            pdf_text("Pacientes", label_style),
            pdf_text("Crianças", label_style),
            pdf_text("Idosos 60+", label_style),
        ],
        [
            pdf_count(stats["total_casas"], number_style),
            pdf_count(stats["casas_vazias"], number_style),
            pdf_count(stats["total_pacientes"], number_style),
            pdf_count(stats["criancas"], number_style),
            pdf_count(stats["idosos"], number_style),
        ],
    ]
    summary_bottom = [
        [
            pdf_text("Gestantes", label_style),
            pdf_text("Homens", label_style),
            pdf_text("Mulheres", label_style),
        ],
        [
            pdf_count(stats["gestantes"], number_style),
            pdf_count(stats["homens"], number_style),
            pdf_count(stats["mulheres"], number_style),
        ],
    ]

    summary_top_table = Table(summary_top, colWidths=[3.4 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm])
    summary_bottom_table = Table(summary_bottom, colWidths=[5.66 * cm, 5.66 * cm, 5.66 * cm])
    summary_style = TableStyle(
        [
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfe2ff")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dee2e6")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fbff")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (0, 1), (-1, 1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
    )
    summary_top_table.setStyle(summary_style)
    summary_bottom_table.setStyle(summary_style)
    resumo_titulo = "Resumo do filtro" if modo_filtro else "Resumo geral"
    elements.extend(
        [
            Paragraph(resumo_titulo, section_style),
            summary_top_table,
            Spacer(1, 4),
            summary_bottom_table,
            Spacer(1, 10),
        ]
    )

    comorbidity_rows = [[pdf_text("Condição", label_style), pdf_text("Total", label_style), pdf_text("Condição", label_style), pdf_text("Total", label_style)]]
    selected_condition_labels = set(filtro_labels)
    selected_condition_cells = []
    selected_condition_style = ParagraphStyle(
        "CondicaoSelecionada",
        parent=normal_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#084298"),
    )
    selected_count_style = ParagraphStyle(
        "TotalCondicaoSelecionada",
        parent=selected_condition_style,
        alignment=1,
    )
    comorbidades = stats["comorbidades"]
    for index in range(0, len(comorbidades), 2):
        left_name, left_total = comorbidades[index]
        right_name = ""
        right_total = ""
        if index + 1 < len(comorbidades):
            right_name, right_total = comorbidades[index + 1]
        row_number = len(comorbidity_rows)
        left_selected = modo_filtro and left_name in selected_condition_labels
        right_selected = modo_filtro and right_name in selected_condition_labels
        if left_selected:
            selected_condition_cells.append((row_number, 0, 1))
        if right_selected:
            selected_condition_cells.append((row_number, 2, 3))
        comorbidity_rows.append(
            [
                pdf_text(left_name, selected_condition_style if left_selected else normal_style),
                pdf_count(left_total, selected_count_style if left_selected else normal_style),
                pdf_text(right_name, selected_condition_style if right_selected else normal_style),
                pdf_count(right_total, selected_count_style if right_selected else normal_style) if right_name else pdf_text("", normal_style),
            ]
        )

    comorbidity_table = Table(comorbidity_rows, repeatRows=1, colWidths=[7 * cm, 1.5 * cm, 7 * cm, 1.5 * cm])
    comorbidity_style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dee2e6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
    ]
    for row_number, start_col, end_col in selected_condition_cells:
        comorbidity_style_commands.extend(
            [
                ("BACKGROUND", (start_col, row_number), (end_col, row_number), colors.HexColor("#fff3cd")),
                ("BOX", (start_col, row_number), (end_col, row_number), 0.5, colors.HexColor("#ffc107")),
            ]
        )
    comorbidity_table.setStyle(TableStyle(comorbidity_style_commands))
    comorbidity_title = "Contagem de comorbidades no filtro" if modo_filtro else "Contagem por comorbidade"
    elements.extend([Paragraph(comorbidity_title, section_style), comorbidity_table, Spacer(1, 12)])

    # Pacientes ainda sem casa vinculada (ex.: recém-importados do e-SUS)
    # ganham seção própria — o relatório NUNCA omite gente do recorte.
    pacientes_sem_casa = pacientes_por_casa.get(None, [])

    estilo_cabecalho_bloco = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f1f6ff")),
            ("SPAN", (0, 1), (1, 1)),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#b6d4fe")),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#084298")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ]
    )

    def montar_cabecalho_bloco(titulo, subtitulo_html, total):
        """Cabeçalho azul de um bloco de pacientes (casa ou 'sem casa')."""
        tabela = Table(
            [
                [Paragraph(titulo, house_style), Paragraph(f"{total} paciente(s)", house_style)],
                [Paragraph(subtitulo_html, normal_style), ""],
            ],
            colWidths=[13.0 * cm, 4.0 * cm],
        )
        tabela.setStyle(estilo_cabecalho_bloco)
        return tabela

    def montar_tabela_pacientes(lista_pacientes):
        """Tabela de pacientes de um bloco. Retorna (tabela, nº de linhas) —
        blocos curtos são mantidos juntos na paginação."""
        dados = [
            [
                pdf_text("Paciente", label_style),
                pdf_text("CPF/CNS", label_style),
                pdf_text("Nascimento", label_style),
                pdf_text("Telefone", label_style),
                pdf_text("Filiação", label_style),
            ]
        ]

        if lista_pacientes:
            for paciente in lista_pacientes:
                filiacao = (
                    f"<b>Pai:</b> {xml_escape(paciente['nome_pai'] or '-')}<br/>"
                    f"<b>Mãe:</b> {xml_escape(paciente['nome_mae'] or '-')}"
                )
                dados.append(
                    [
                        pdf_text(paciente["nome"], normal_style),
                        pdf_text(formatar_cpf_ou_cns(paciente["cpf"]), normal_style),
                        pdf_text(formatar_data_br(paciente["data_nascimento"]), normal_style),
                        pdf_text(formatar_telefone(paciente["telefone"]), normal_style),
                        Paragraph(filiacao, normal_style),
                    ]
                )
                if paciente["observacao"]:
                    dados.append(
                        [
                            Paragraph(f"<b>Observação:</b> {xml_escape(paciente['observacao'] or '')}", normal_style),
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
                condicoes = listar_condicoes(paciente["condicoes_saude"])
                if condicoes:
                    dados.append(
                        [
                            Paragraph(f"<b>Condições de saúde:</b> {', '.join(condicoes)}", normal_style),
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
        else:
            dados.append([Paragraph("Nenhum paciente cadastrado", empty_style), "", "", "", ""])

        tabela = Table(dados, repeatRows=1, colWidths=[3.6 * cm, 2.8 * cm, 2.2 * cm, 2.7 * cm, 5.7 * cm])
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#212529")),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#dee2e6")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        for row_index, row in enumerate(dados):
            if row_index > 0 and row[1:] == ["", "", "", ""]:
                tabela.setStyle(
                    TableStyle(
                        [
                            ("SPAN", (0, row_index), (-1, row_index)),
                            ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#f1f6ff")),
                        ]
                    )
                )
        return tabela, len(dados)

    def adicionar_bloco(intro_elements, cabecalho, tabela, linhas):
        intro = intro_elements + [cabecalho, Spacer(1, 6)]
        if linhas <= 8:
            elements.append(KeepTogether(intro + [tabela, Spacer(1, 14)]))
        else:
            elements.append(KeepTogether(intro))
            elements.append(tabela)
            elements.append(Spacer(1, 14))

    if not casas and not pacientes_sem_casa:
        if filtro_sem_selecao:
            empty_message = "Nenhum grupo ou comorbidade selecionado para exportação."
        elif filtro_ativo:
            empty_message = "Nenhum paciente encontrado para o recorte selecionado."
        else:
            empty_message = "Nenhuma casa ou paciente cadastrado."
        elements.append(
            KeepTogether(
                [
                    Paragraph("Pacientes por quadra e casa", section_style),
                    Paragraph(empty_message, normal_style),
                ]
            )
        )

    quadra_atual = object()
    primeiro_bloco = True
    for casa in casas:
        casa_elements = []
        quadra_label = f"Quadra Nº {casa['numero_quadra']}" if casa["numero_quadra"] else "Sem quadra"

        if primeiro_bloco:
            casa_elements.append(Paragraph("Pacientes por quadra e casa", section_style))

        if quadra_label != quadra_atual:
            casa_elements.append(Paragraph(quadra_label, block_style))
            quadra_atual = quadra_label

        lista_pacientes = pacientes_por_casa.get(casa["id"], [])

        tipo_casa = tipo_imovel_de(casa)
        endereco_pdf = f"<b>Endereço:</b> {xml_escape(casa['endereco'] or '')}"
        if tipo_casa != "domicilio":
            endereco_pdf += f" &nbsp;|&nbsp; <b>Tipo:</b> {xml_escape(tipo_imovel_label(tipo_casa))}"

        cabecalho = montar_cabecalho_bloco(f"Casa Nº {casa['numero_casa']}", endereco_pdf, len(lista_pacientes))
        tabela, linhas = montar_tabela_pacientes(lista_pacientes)
        adicionar_bloco(casa_elements, cabecalho, tabela, linhas)
        primeiro_bloco = False

    if pacientes_sem_casa:
        secao_elements = []
        if primeiro_bloco:
            secao_elements.append(Paragraph("Pacientes por quadra e casa", section_style))
        secao_elements.append(Paragraph("Sem casa definida", block_style))
        cabecalho = montar_cabecalho_bloco(
            "Pacientes sem casa vinculada",
            "<b>Endereço:</b> ainda não definido no sistema — vincule pela página Pacientes",
            len(pacientes_sem_casa),
        )
        tabela, linhas = montar_tabela_pacientes(pacientes_sem_casa)
        adicionar_bloco(secao_elements, cabecalho, tabela, linhas)

    doc.build(elements, onFirstPage=adicionar_cabecalho_rodape, onLaterPages=adicionar_cabecalho_rodape)
    buffer.seek(0)
    # O nome do arquivo declara o recorte — vários PDFs baixados não viram
    # uma pilha de "relatorio (3).pdf" indistinguíveis.
    partes_nome = list(grupos_selecionados)
    if condicoes_selecionadas:
        partes_nome.append("comorbidades")
    if partes_nome:
        nome_pdf = "relatorio_" + "_".join(partes_nome) + ".pdf"
    else:
        nome_pdf = "relatorio_pacientes_por_casa.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_pdf,
        mimetype="application/pdf",
    )


@app.route("/exportar/perfil-epidemiologico", methods=["GET", "POST"])
@login_required
def exportar_perfil_epidemiologico():
    # POST salva a identificação do cabeçalho antes de exportar; GET usa a
    # salva. Nas duas formas o download é o mesmo documento.
    if request.method == "POST":
        set_preferencia("acs_nome", request.form.get("acs_nome", "").strip()[:120])
        set_preferencia("acs_microarea", request.form.get("acs_microarea", "").strip()[:20])
    acs_nome = get_preferencia("acs_nome")
    acs_microarea = get_preferencia("acs_microarea")

    # Mês corrente sempre ao vivo (e persistido); anteriores vêm do histórico,
    # com mês sem uso herdando o retrato anterior. Registrar ANTES de montar a
    # série: no primeiro uso é este retrato que ancora o início dela.
    perfil_atual = registrar_perfil_mensal()
    serie_completa = meses_da_serie_perfil()
    # A herança é montada sobre a série inteira e o recorte é só de exibição:
    # se o 1º mês da janela não tem registro, ele herda de antes da janela.
    perfis = montar_serie_perfis(serie_completa, perfil_atual)
    meses = serie_completa[-PERFIL_MESES_NO_PDF:]

    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")
    largura_pagina = landscape(A4)[0]

    def adicionar_rodape(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#6c757d"))
        canvas.drawString(1.2 * cm, 0.8 * cm, "Saúde Simples")
        canvas.drawRightString(largura_pagina - 1.2 * cm, 0.8 * cm, f"Página {doc.page}")
        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.3 * cm,
    )
    styles = getSampleStyleSheet()
    overline_style = ParagraphStyle(
        "OverlinePerfil",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#0b5ed7"),
        spaceAfter=3,
    )
    title_style = ParagraphStyle(
        "TituloPerfil",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=19,
        leading=22,
        textColor=colors.HexColor("#212529"),
        alignment=0,
        spaceAfter=0,
    )
    meta_style = ParagraphStyle(
        "MetaPerfil",
        parent=styles["BodyText"],
        fontSize=8,
        leading=12,
        textColor=colors.HexColor("#6c757d"),
        alignment=2,
    )
    id_style = ParagraphStyle(
        "IdentificacaoPerfil",
        parent=styles["BodyText"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#495057"),
        spaceBefore=4,
    )
    label_style = ParagraphStyle(
        "RotuloPerfil",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#334155"),
    )

    header_table = Table(
        [
            [
                [
                    Paragraph("SAÚDE SIMPLES", overline_style),
                    Paragraph("Perfil Epidemiológico", title_style),
                ],
                Paragraph(f"Gerado em {data_geracao}", meta_style),
            ]
        ],
        colWidths=[17 * cm, 10.3 * cm],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (0, 0), "TOP"),
                ("VALIGN", (1, 0), (1, 0), "BOTTOM"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    header_rule = Table([["", ""]], colWidths=[2.6 * cm, 24.7 * cm], rowHeights=[2])
    header_rule.setStyle(
        TableStyle(
            [
                ("LINEBELOW", (0, 0), (0, 0), 1.6, colors.HexColor("#0b5ed7")),
                ("LINEBELOW", (1, 0), (1, 0), 0.5, colors.HexColor("#dee2e6")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    elements = [header_table, Spacer(1, 6), header_rule]
    identificacao = []
    if acs_nome:
        identificacao.append(f"<b>ACS:</b> {xml_escape(acs_nome)}")
    if acs_microarea:
        identificacao.append(f"<b>Microárea:</b> {xml_escape(acs_microarea)}")
    if identificacao:
        elements.append(Paragraph(" &#183; ".join(identificacao), id_style))
    elements.append(Spacer(1, 8))

    # --- Grade: coluna de rótulos + 2 subcolunas (F/M) por mês -------------
    # A série completa é fatiada em blocos de 7 meses (a largura da folha);
    # cada bloco vira uma tabela e o documento pagina sozinho — o histórico
    # cresce sem limite de janela.
    MESES_POR_BLOCO = 7
    # Largura de subcoluna fixa (a do bloco cheio): bloco parcial sai com as
    # mesmas proporções, alinhado à esquerda, em vez de esticado.
    largura_mes = (27.3 - 3.9) / (MESES_POR_BLOCO * 2)

    def pad(valor):
        return f"{int(valor):02d}"

    def montar_bloco(bloco):
        grade = [[Paragraph("Mês/Ano", label_style)]]
        for ano_mes in bloco:
            grade[0].extend([rotulo_mes(ano_mes), ""])

        spans_meia_linha = []  # células que ocupam as 2 subcolunas do mês
        celulas_sem_dado = []  # células "-----": mês sem retrato registrado

        linha_familias = [Paragraph("Famílias cadastradas", label_style)]
        linha_genero = [Paragraph("Gênero", label_style)]
        for indice_mes, ano_mes in enumerate(bloco):
            dados = perfis[ano_mes]
            valor = dados.get("familias") if dados else None
            if valor is None:
                celulas_sem_dado.append((1 + indice_mes * 2, 1))
            linha_familias.extend(["-----" if valor is None else pad(valor), ""])
            linha_genero.extend(["F", "M"])
        grade.extend([linha_familias, linha_genero])

        for rotulo, chave, tipo in PERFIL_LINHAS:
            linha = [Paragraph(rotulo, label_style)]
            indice_linha = len(grade)
            for indice_mes, ano_mes in enumerate(bloco):
                dados = perfis[ano_mes]
                valor = dados.get(chave) if dados else None
                coluna = 1 + indice_mes * 2
                if valor is None:
                    linha.extend(["-----", ""])
                    spans_meia_linha.append((coluna, indice_linha))
                    celulas_sem_dado.append((coluna, indice_linha))
                elif tipo == "span":
                    linha.extend([pad(valor), ""])
                    spans_meia_linha.append((coluna, indice_linha))
                else:
                    linha.extend([pad(valor.get("f", 0)), pad(valor.get("m", 0))])
            grade.append(linha)

        tabela = Table(
            grade,
            colWidths=[3.9 * cm] + [largura_mes * cm] * (len(bloco) * 2),
            repeatRows=1,
            hAlign="LEFT",
        )
        # Comandos do geral para o específico — os últimos vencem nas células
        # onde há sobreposição.
        estilo = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#adb5bd")),
            ("FONTNAME", (1, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (1, 0), (-1, -1), 8),
            ("TEXTCOLOR", (1, 1), (-1, -1), colors.HexColor("#212529")),
            ("FONTNAME", (1, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (1, 0), (-1, 0), colors.HexColor("#212529")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef4fb")),
            ("FONTNAME", (1, 2), (-1, 2), "Helvetica-Bold"),
            ("TEXTCOLOR", (1, 2), (-1, 2), colors.HexColor("#495057")),
            ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#f8fafc")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ]
        # Zebra discreta nas linhas de dados (antes dos realces por célula).
        for linha in range(3, len(grade)):
            if (linha - 3) % 2 == 1:
                estilo.append(("BACKGROUND", (0, linha), (-1, linha), colors.HexColor("#fbfdff")))
        # Cabeçalho dos meses e linha de famílias: cada mês ocupa as 2 subcolunas.
        for indice_mes in range(len(bloco)):
            coluna = 1 + indice_mes * 2
            estilo.append(("SPAN", (coluna, 0), (coluna + 1, 0)))
            estilo.append(("SPAN", (coluna, 1), (coluna + 1, 1)))
        for coluna, linha in spans_meia_linha:
            estilo.append(("SPAN", (coluna, linha), (coluna + 1, linha)))
        for coluna, linha in celulas_sem_dado:
            estilo.append(("TEXTCOLOR", (coluna, linha), (coluna, linha), colors.HexColor("#94a3b8")))
        tabela.setStyle(TableStyle(estilo))
        return tabela

    for inicio_bloco in range(0, len(meses), MESES_POR_BLOCO):
        bloco = meses[inicio_bloco:inicio_bloco + MESES_POR_BLOCO]
        # KeepTogether: um bloco nunca quebra no meio entre páginas.
        elements.append(KeepTogether(montar_bloco(bloco)))
        elements.append(Spacer(1, 12))

    legenda_style = ParagraphStyle(
        "LegendaPerfil",
        parent=styles["BodyText"],
        fontName="Helvetica-Oblique",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor("#6c757d"),
        spaceBefore=6,
    )
    elements.append(
        Paragraph(
            "O documento apresenta os últimos 12 meses — o histórico completo fica guardado no "
            "sistema e sai integral na planilha (XLSX). O retrato de cada mês é gravado "
            "automaticamente durante o uso do sistema; mês em que o sistema não foi usado mantém o "
            "retrato do mês anterior (o cadastro não mudou nesse período). "
            "“-----” marca apenas indicador sem registro na época.",
            legenda_style,
        )
    )

    doc.build(elements, onFirstPage=adicionar_rodape, onLaterPages=adicionar_rodape)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"perfil_epidemiologico_{meses[-1]}.pdf",
        mimetype="application/pdf",
    )


@app.route("/exportar/perfil-epidemiologico.xlsx")
@login_required
def exportar_perfil_epidemiologico_xlsx():
    """Histórico completo do perfil em planilha XLSX estilizada.

    Meses como linhas — o formato analítico: ordenável, filtrável, pronto
    para gráfico e tabela dinâmica. Indicadores em colunas com F/M agrupados.
    A coluna Origem preserva a auditoria que o PDF não mostra: qual mês foi
    registrado de fato, qual herdou o anterior e qual é o retrato ao vivo.
    """
    perfil_atual = registrar_perfil_mensal()
    meses = meses_da_serie_perfil()
    perfis = montar_serie_perfis(meses, perfil_atual)
    meses_registrados = set(carregar_perfis_mensais(meses[:-1]))
    acs_nome = get_preferencia("acs_nome")
    acs_microarea = get_preferencia("acs_microarea")

    AZUL = "0B5ED7"
    ESCURO = "212529"
    CINZA = "64748B"
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"
    ws.sheet_view.showGridLines = False

    fio = Side(style="thin", color="DEE2E6")
    borda = Border(left=fio, right=fio, top=fio, bottom=fio)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fonte_cabecalho = Font(bold=True, size=9, color="FFFFFF")
    fill_cabecalho = PatternFill("solid", fgColor=AZUL)
    fonte_sub = Font(bold=True, size=8, color="334155")
    fill_sub = PatternFill("solid", fgColor="EEF4FB")
    fill_zebra = PatternFill("solid", fgColor="F8FAFC")

    # --- Título ------------------------------------------------------------
    ws.cell(row=1, column=1, value="SAÚDE SIMPLES").font = Font(bold=True, size=8, color=AZUL)
    ws.cell(row=2, column=1, value="Perfil Epidemiológico — Histórico Mensal").font = Font(
        bold=True, size=14, color=ESCURO
    )
    identificacao = []
    if acs_nome:
        identificacao.append(f"ACS: {acs_nome}")
    if acs_microarea:
        identificacao.append(f"Microárea: {acs_microarea}")
    identificacao.append(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    ws.cell(row=3, column=1, value=" · ".join(identificacao)).font = Font(size=9, color=CINZA)
    ws.row_dimensions[2].height = 20

    # --- Cabeçalho em 2 linhas: indicador em cima, F/M embaixo -------------
    LINHA_CAB, LINHA_SUB, LINHA_DADOS = 5, 6, 7

    def cabecalho(linha, coluna, valor, sub=False):
        c = ws.cell(row=linha, column=coluna, value=valor)
        c.font = fonte_sub if sub else fonte_cabecalho
        c.fill = fill_sub if sub else fill_cabecalho
        c.alignment = centro
        return c

    coluna = 1
    for titulo in ("Mês/Ano", "Origem do retrato", "Famílias cadastradas"):
        ws.merge_cells(start_row=LINHA_CAB, start_column=coluna, end_row=LINHA_SUB, end_column=coluna)
        cabecalho(LINHA_CAB, coluna, titulo)
        coluna += 1
    colunas_indicadores = []  # (chave, tipo, primeira coluna)
    for rotulo, chave, tipo in PERFIL_LINHAS:
        colunas_indicadores.append((chave, tipo, coluna))
        if tipo == "span":
            ws.merge_cells(start_row=LINHA_CAB, start_column=coluna, end_row=LINHA_SUB, end_column=coluna)
            cabecalho(LINHA_CAB, coluna, rotulo)
            coluna += 1
        else:
            ws.merge_cells(start_row=LINHA_CAB, start_column=coluna, end_row=LINHA_CAB, end_column=coluna + 1)
            cabecalho(LINHA_CAB, coluna, rotulo)
            cabecalho(LINHA_SUB, coluna, "F", sub=True)
            cabecalho(LINHA_SUB, coluna + 1, "M", sub=True)
            coluna += 2
    total_colunas = coluna - 1
    ws.row_dimensions[LINHA_CAB].height = 30
    ws.row_dimensions[LINHA_SUB].height = 14

    # --- Dados: um mês por linha -------------------------------------------
    fonte_valor = Font(size=9, color=ESCURO)
    fontes_origem = {
        "registrado": Font(size=8.5, color="334155"),
        "herdado do mês anterior": Font(size=8.5, italic=True, color=CINZA),
        "mês corrente (ao vivo)": Font(size=8.5, bold=True, color=AZUL),
    }

    def valor_numerico(linha, coluna_alvo, valor, zebra):
        c = ws.cell(row=linha, column=coluna_alvo)
        if valor is not None:
            c.value = int(valor)
            c.number_format = "00"
        c.font = fonte_valor
        c.alignment = centro
        if zebra:
            c.fill = fill_zebra

    for indice, ano_mes in enumerate(meses):
        linha = LINHA_DADOS + indice
        dados = perfis[ano_mes]
        zebra = indice % 2 == 1
        if ano_mes == meses[-1]:
            origem = "mês corrente (ao vivo)"
        elif ano_mes in meses_registrados:
            origem = "registrado"
        else:
            origem = "herdado do mês anterior"

        celula_mes = ws.cell(row=linha, column=1, value=rotulo_mes(ano_mes))
        celula_mes.font = Font(bold=True, size=9, color=ESCURO)
        celula_mes.alignment = centro
        celula_origem = ws.cell(row=linha, column=2, value=origem)
        celula_origem.font = fontes_origem[origem]
        celula_origem.alignment = Alignment(horizontal="left", vertical="center")
        if zebra:
            celula_mes.fill = fill_zebra
            celula_origem.fill = fill_zebra

        valor_numerico(linha, 3, dados.get("familias") if dados else None, zebra)
        for chave, tipo, coluna_inicial in colunas_indicadores:
            valor = dados.get(chave) if dados else None
            if tipo == "span":
                valor_numerico(linha, coluna_inicial, valor, zebra)
            else:
                valor_numerico(linha, coluna_inicial, valor.get("f", 0) if valor else None, zebra)
                valor_numerico(linha, coluna_inicial + 1, valor.get("m", 0) if valor else None, zebra)

    # Bordas em toda a área da tabela (inclusive células mescladas).
    ultima_linha = LINHA_DADOS + len(meses) - 1
    for linha in range(LINHA_CAB, ultima_linha + 1):
        for coluna_alvo in range(1, total_colunas + 1):
            ws.cell(row=linha, column=coluna_alvo).border = borda

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 23
    ws.column_dimensions["C"].width = 11
    for coluna_alvo in range(4, total_colunas + 1):
        ws.column_dimensions[get_column_letter(coluna_alvo)].width = 7
    # Indicador de coluna única (ex.: Gestantes) precisa caber o rótulo inteiro.
    for _, tipo, coluna_inicial in colunas_indicadores:
        if tipo == "span":
            ws.column_dimensions[get_column_letter(coluna_inicial)].width = 11
    # Cabeçalho e identificação do mês sempre à vista ao rolar.
    ws.freeze_panes = f"C{LINHA_DADOS}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"perfil_epidemiologico_historico_{meses[-1]}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/banco")
@login_required
def banco():
    conn = get_db_connection()
    contagens = {
        "quadras": conn.execute("SELECT COUNT(*) AS c FROM quadras").fetchone()["c"],
        "casas": conn.execute("SELECT COUNT(*) AS c FROM casas").fetchone()["c"],
        "pacientes": conn.execute("SELECT COUNT(*) AS c FROM pacientes").fetchone()["c"],
    }
    conn.close()

    try:
        tamanho_banco = os.path.getsize(DATABASE)
    except OSError:
        tamanho_banco = 0

    return render_template(
        "banco.html",
        contagens=contagens,
        tamanho_banco=tamanho_banco,
        backups=listar_backups(),
    )


@app.route("/banco/backup", methods=["POST"])
@login_required
def criar_backup_manual():
    try:
        criar_backup("manual")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup manual: %s", exc)
        flash("Não foi possível criar o backup. Tente novamente.", "danger")
        return redirect(url_for("banco"))

    flash("Backup criado com sucesso.", "success")
    return redirect(url_for("banco"))


@app.route("/banco/exportar")
@login_required
def exportar_banco():
    # Snapshot consistente via API de backup do SQLite — nunca o arquivo cru,
    # que poderia estar no meio de uma transação WAL.
    descritor, caminho_tmp = tempfile.mkstemp(suffix=".db")
    os.close(descritor)
    try:
        exportar_snapshot(caminho_tmp)
        with open(caminho_tmp, "rb") as arquivo:
            conteudo = BytesIO(arquivo.read())
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao exportar banco: %s", exc)
        flash("Não foi possível exportar o banco. Tente novamente.", "danger")
        return redirect(url_for("banco"))
    finally:
        try:
            os.remove(caminho_tmp)
        except OSError:
            pass

    conteudo.seek(0)
    logger.info("Banco exportado a partir de %s", request.remote_addr)
    nome = f"saude_simples_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return send_file(
        conteudo,
        as_attachment=True,
        download_name=nome,
        mimetype="application/x-sqlite3",
    )


@app.route("/banco/backup/<nome>/baixar")
@login_required
def baixar_backup(nome):
    try:
        caminho = caminho_backup(nome)
    except ValueError:
        flash("Backup não encontrado.", "warning")
        return redirect(url_for("banco"))

    return send_file(
        caminho,
        as_attachment=True,
        download_name=nome,
        mimetype="application/x-sqlite3",
    )


@app.route("/banco/restaurar", methods=["POST"])
@login_required
def restaurar_banco():
    nome = request.form.get("nome", "")
    try:
        caminho_backup(nome)
    except ValueError:
        flash("Backup não encontrado.", "warning")
        return redirect(url_for("banco"))

    try:
        criar_backup("antes_restaurar")
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao criar backup antes de restaurar: %s", exc)
        flash("Não foi possível criar backup de segurança. Restauração cancelada.", "danger")
        return redirect(url_for("banco"))

    # A senha de acesso ATUAL sobrevive à restauração — voltar os dados no
    # tempo nunca tranca o operador para fora. Para recuperar senha esquecida,
    # o caminho é o resetar_senha.py.
    senha_atual = get_senha_hash()
    try:
        restaurar_backup(nome)
    except (sqlite3.Error, OSError) as exc:
        logger.error("Falha ao restaurar backup %s: %s", nome, exc)
        flash("Não foi possível restaurar o backup.", "danger")
        return redirect(url_for("banco"))
    if senha_atual:
        set_senha_hash(senha_atual)

    logger.info("Backup %s restaurado a partir de %s", nome, request.remote_addr)
    flash("Backup restaurado com sucesso.", "success")
    return redirect(url_for("banco"))


@app.route("/banco/importar", methods=["POST"])
@login_required
def importar_banco_view():
    arquivo = request.files.get("arquivo")
    if arquivo is None or not arquivo.filename:
        flash("Selecione o arquivo .db exportado do Saúde Simples.", "danger")
        return redirect(url_for("banco"))

    descritor, caminho_tmp = tempfile.mkstemp(suffix=".db")
    os.close(descritor)
    try:
        arquivo.save(caminho_tmp)

        try:
            criar_backup("antes_importar")
        except (sqlite3.Error, OSError) as exc:
            logger.error("Falha ao criar backup antes de importar: %s", exc)
            flash("Não foi possível criar backup de segurança. Importação cancelada.", "danger")
            return redirect(url_for("banco"))

        try:
            importar_banco(caminho_tmp)
        except BancoInvalido as exc:
            logger.warning("Importação de banco rejeitada (%s): %s", request.remote_addr, exc)
            flash(str(exc), "danger")
            return redirect(url_for("banco"))
        except (sqlite3.Error, OSError) as exc:
            logger.error("Falha ao importar banco: %s", exc)
            flash("Não foi possível importar o banco. O banco atual foi preservado.", "danger")
            return redirect(url_for("banco"))
    finally:
        try:
            os.remove(caminho_tmp)
        except OSError:
            pass

    logger.info("Banco importado a partir de %s", request.remote_addr)
    flash("Banco importado com sucesso. A senha de acesso atual foi mantida.", "success")
    return redirect(url_for("banco"))


def preparar_primeiro_boot():
    """Prepara o sistema para servir: banco criado/migrado, senha garantida
    (ou código de configuração inicial no terminal) e backup de inicialização.
    Chamado pelo run.py — a entrada canônica do servidor."""
    init_db()
    # Honra a retenção da lixeira mesmo que a página nunca seja aberta.
    conn = get_db_connection()
    purgar_lixeira_expirada(conn)
    conn.commit()
    conn.close()
    if not garantir_senha_inicial(BOOTSTRAP_PASSWORD_HASH):
        codigo_setup = ensure_setup_token()
        # print, não logger: este é o canal de entrega do código — precisa
        # saltar aos olhos de quem acabou de iniciar o servidor.
        print()
        print("=" * 62)
        print("  PRIMEIRO ACESSO — nenhuma senha configurada ainda.")
        print()
        print(f"  Código de segurança:  {codigo_setup}")
        print()
        print("  Abra o sistema no navegador e informe este código na tela")
        print("  de configuração inicial para definir a senha de acesso.")
        print("=" * 62)
        print()
    try:
        criar_backup("inicializacao")
    except (sqlite3.Error, OSError) as exc:
        logger.warning("Falha ao criar backup de inicialização: %s", exc)


if __name__ == "__main__":
    # Entrada canônica é o run.py (rede WiFi local + TLS automático).
    # `python app.py` continua funcionando: delega para lá.
    import runpy

    runpy.run_path(os.path.join(BASE_DIR, "run.py"), run_name="__main__")
