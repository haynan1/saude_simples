"""Perfil Epidemiológico — série mensal automatizada e exportação em PDF.

O retrato do mês corrente é sempre calculado ao vivo; meses anteriores vêm da
tabela perfil_mensal (registrada automaticamente durante o uso). Mês sem
registro aparece como "-----" no documento.
"""
import json
from datetime import datetime

import db
from tests.conftest import criar_casa, criar_paciente, criar_quadra, texto_pdf


def _ano_mes_atual():
    return datetime.now().strftime("%Y-%m")


def _seed_territorio(client):
    criar_quadra(client)
    criar_casa(client)
    criar_paciente(
        client,
        nome="MULHER HIPERTENSA FUMANTE",
        cpf="11111111111",
        sexo="Feminino",
        data_nascimento="1950-05-10",  # idosa
        condicoes_saude=["hipertensao", "fumante", "bolsa_familia"],
    )
    criar_paciente(
        client,
        nome="HOMEM DIABETICO",
        cpf="22222222222",
        sexo="Masculino",
        data_nascimento="1990-05-10",
        condicoes_saude=["diabetes", "deficiencia"],
    )
    criar_paciente(
        client,
        nome="BEBE DA CASA",
        cpf="33333333333",
        sexo="Masculino",
        data_nascimento=datetime.now().strftime("%Y-01-10"),  # 0 a 2 anos
    )


def test_calculo_do_perfil_separa_por_sexo(logged_client):
    import app as app_module

    _seed_territorio(logged_client)
    perfil = app_module.calcular_perfil_epidemiologico()

    assert perfil["familias"] == 1
    assert perfil["pessoas"] == {"f": 1, "m": 2}
    assert perfil["hipertensos"] == {"f": 1, "m": 0}
    assert perfil["diabeticos"] == {"f": 0, "m": 1}
    assert perfil["idosos"] == {"f": 1, "m": 0}
    assert perfil["criancas_0_2"] == {"f": 0, "m": 1}
    # Pedido explícito: tabagistas separado por sexo (não um total único).
    assert perfil["tabagistas"] == {"f": 1, "m": 0}
    assert perfil["deficientes"] == {"f": 0, "m": 1}
    assert perfil["bolsa_familia"] == {"f": 1, "m": 0}


def test_exportacao_gera_pdf_com_mes_corrente(logged_client):
    _seed_territorio(logged_client)
    resp = logged_client.get("/exportar/perfil-epidemiologico")
    assert resp.status_code == 200
    assert resp.mimetype == "application/pdf"
    texto = texto_pdf(resp.data)
    assert "Perfil Epidemiol" in texto
    assert "Tabagistas" in texto
    # Meses sem retrato registrado aparecem como tracinhos.
    assert "-----" in texto


def test_exportacao_persiste_retrato_do_mes(logged_client):
    _seed_territorio(logged_client)
    logged_client.get("/exportar/perfil-epidemiologico")
    conn = db.get_db_connection()
    row = conn.execute(
        "SELECT dados FROM perfil_mensal WHERE ano_mes = ?", (_ano_mes_atual(),)
    ).fetchone()
    conn.close()
    assert row is not None
    assert json.loads(row["dados"])["pessoas"] == {"f": 1, "m": 2}


def _mes_anterior(ano_mes, quantos=1):
    ano, mes = int(ano_mes[:4]), int(ano_mes[5:7])
    mes -= quantos
    while mes <= 0:
        ano, mes = ano - 1, mes + 12
    return f"{ano:04d}-{mes:02d}"


def test_historico_congelado_entra_no_documento(logged_client):
    _seed_territorio(logged_client)
    # Retrato "congelado" de um mês anterior, com valor impossível de vir do
    # cálculo ao vivo — se aparece, veio do histórico.
    mes_anterior = _mes_anterior(_ano_mes_atual())
    db.salvar_perfil_mensal(
        mes_anterior,
        json.dumps({"familias": 87, "pessoas": {"f": 43, "m": 44}}),
    )
    texto = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    assert "87" in texto
    assert "43" in texto
    # Indicador ausente no retrato antigo (esquema evoluiu): vira "-----",
    # nunca zero inventado.
    assert "-----" in texto


def test_serie_nasce_no_primeiro_mes_alimentado(logged_client):
    """Sem histórico, o documento mostra só o mês corrente — nada de meses
    vazios antes do sistema existir."""
    import app as app_module

    _seed_territorio(logged_client)
    texto = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    atual = _ano_mes_atual()
    assert app_module.rotulo_mes(atual) in texto
    assert app_module.rotulo_mes(_mes_anterior(atual)) not in texto


def test_serie_cresce_a_partir_da_ancora(logged_client):
    """Com um retrato 3 meses atrás, a série vai da âncora ao mês corrente —
    meses sem registro herdam o retrato da âncora; indicadores que não
    existiam nele saem como '-----'; antes da âncora, nada."""
    import app as app_module

    _seed_territorio(logged_client)
    atual = _ano_mes_atual()
    ancora = _mes_anterior(atual, 3)
    db.salvar_perfil_mensal(ancora, json.dumps({"familias": 87}))

    texto = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    for quantos in range(4):  # âncora, os 2 meses herdados e o mês corrente
        assert app_module.rotulo_mes(_mes_anterior(atual, quantos)) in texto
    assert app_module.rotulo_mes(_mes_anterior(atual, 4)) not in texto
    assert "-----" in texto  # indicadores ausentes no retrato da âncora


def _retrato_completo(familias):
    """Retrato com todos os indicadores preenchidos, como o gravado pelo
    registro automático."""
    import app as app_module

    dados = {"familias": familias}
    for _, chave, tipo in app_module.PERFIL_LINHAS:
        dados[chave] = 5 if tipo == "span" else {"f": 2, "m": 3}
    return dados


def test_mes_sem_uso_herda_o_retrato_anterior(logged_client):
    """O relatório é entregue todo mês: sem uso do sistema no mês, vale o
    último retrato registrado — o cadastro não mudou nesse período."""
    import app as app_module

    _seed_territorio(logged_client)
    atual = _ano_mes_atual()
    ancora = _mes_anterior(atual, 2)
    db.salvar_perfil_mensal(ancora, json.dumps(_retrato_completo(87)))

    meses = app_module.meses_da_serie_perfil()
    perfis = app_module.montar_serie_perfis(
        meses, app_module.calcular_perfil_epidemiologico()
    )
    # O mês sem registro herda o retrato da âncora, valor a valor.
    assert perfis[_mes_anterior(atual, 1)] == perfis[ancora]

    texto = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    assert "87" in texto
    # Série contínua e completa: nenhum tracinho na TABELA — documento
    # entregável. (A única ocorrência de "-----" é a citação da legenda.)
    assert texto.count("-----") == 1


def test_meses_da_serie_atravessa_virada_de_ano(logged_client):
    from datetime import datetime as dt

    import app as app_module

    db.salvar_perfil_mensal("2025-11", json.dumps({"familias": 1}))
    meses = app_module.meses_da_serie_perfil(referencia=dt(2026, 2, 15))
    assert meses == ["2025-11", "2025-12", "2026-01", "2026-02"]


def test_navegacao_registra_retrato_do_dia(logged_client):
    """O hook diário fotografa o mês corrente sem o operador pedir."""
    logged_client.get("/")
    conn = db.get_db_connection()
    row = conn.execute(
        "SELECT dados FROM perfil_mensal WHERE ano_mes = ?", (_ano_mes_atual(),)
    ).fetchone()
    conn.close()
    assert row is not None


def test_post_salva_identificacao_do_acs(logged_client):
    resp = logged_client.post(
        "/exportar/perfil-epidemiologico",
        data={"acs_nome": "Maria da Silva", "acs_microarea": "13"},
    )
    assert resp.status_code == 200
    assert db.get_preferencia("acs_nome") == "Maria da Silva"
    assert db.get_preferencia("acs_microarea") == "13"
    # O cabeçalho do documento carrega a identificação salva.
    texto = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    assert "Maria da Silva" in texto
    # E a página Exportar reapresenta os valores salvos.
    pagina = logged_client.get("/exportar")
    assert b"Maria da Silva" in pagina.data


def test_xlsx_do_historico_com_origem_de_cada_mes(logged_client):
    """A planilha entrega a série completa e a auditoria que o PDF não
    mostra: registrado × herdado × ao vivo."""
    from io import BytesIO

    from openpyxl import load_workbook

    _seed_territorio(logged_client)
    atual = _ano_mes_atual()
    ancora = _mes_anterior(atual, 2)
    db.salvar_perfil_mensal(ancora, json.dumps(_retrato_completo(87)))

    resp = logged_client.get("/exportar/perfil-epidemiologico.xlsx")
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    ws = load_workbook(BytesIO(resp.data)).active
    assert ws.cell(row=5, column=1).value == "Mês/Ano"
    assert ws.cell(row=5, column=2).value == "Origem do retrato"
    assert ws.cell(row=5, column=3).value == "Famílias cadastradas"
    # Subcabeçalho F/M na primeira coluna de indicador por sexo.
    assert ws.cell(row=6, column=4).value == "F"
    assert ws.cell(row=6, column=5).value == "M"
    # âncora + mês herdado + mês corrente = 3 linhas de dados (7, 8 e 9).
    import app as app_module

    assert ws.cell(row=7, column=1).value == app_module.rotulo_mes(ancora)
    assert ws.cell(row=7, column=2).value == "registrado"
    assert ws.cell(row=7, column=3).value == 87
    assert ws.cell(row=8, column=2).value == "herdado do mês anterior"
    assert ws.cell(row=8, column=3).value == 87
    assert ws.cell(row=9, column=1).value == app_module.rotulo_mes(atual)
    assert ws.cell(row=9, column=2).value == "mês corrente (ao vivo)"
    assert ws.cell(row=10, column=1).value is None


def test_pdf_janela_de_12_meses_e_xlsx_integral(logged_client):
    """O PDF é o entregável (últimos 12 meses); o XLSX é o arquivo (tudo).
    Mês que 'sai' do PDF permanece no banco e na planilha — e o 1º mês da
    janela, quando não registrado, herda de antes da janela."""
    from io import BytesIO

    from openpyxl import load_workbook

    import app as app_module

    _seed_territorio(logged_client)
    atual = _ano_mes_atual()
    # Âncora 13 meses atrás: série completa tem 14 meses; só a âncora tem
    # registro — a janela inteira herda dela, atravessando a borda do recorte.
    ancora = _mes_anterior(atual, 13)
    db.salvar_perfil_mensal(ancora, json.dumps(_retrato_completo(87)))

    pdf = texto_pdf(logged_client.get("/exportar/perfil-epidemiologico").data)
    for quantos in range(12):  # do mês corrente até 11 atrás: dentro do PDF
        assert app_module.rotulo_mes(_mes_anterior(atual, quantos)) in pdf
    assert app_module.rotulo_mes(_mes_anterior(atual, 12)) not in pdf
    assert app_module.rotulo_mes(ancora) not in pdf
    # Herança atravessou a borda: valores da âncora presentes, tabela sem
    # tracinho (a única ocorrência de "-----" é a citação da legenda).
    assert "87" in pdf
    assert pdf.count("-----") == 1

    resp = logged_client.get("/exportar/perfil-epidemiologico.xlsx")
    ws = load_workbook(BytesIO(resp.data)).active
    # XLSX integral: 14 linhas — âncora, 12 herdadas e o mês corrente.
    assert ws.cell(row=7, column=1).value == app_module.rotulo_mes(ancora)
    assert ws.cell(row=7 + 13, column=1).value == app_module.rotulo_mes(atual)
    assert ws.cell(row=7 + 14, column=1).value is None


def test_novas_condicoes_no_cadastro(logged_client):
    pagina = logged_client.get("/pacientes/novo")
    assert "Pessoa com deficiência".encode() in pagina.data
    assert "Recebe Bolsa Família".encode() in pagina.data
